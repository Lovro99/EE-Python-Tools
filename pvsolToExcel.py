"""
PVSOL -> Excel ekstraktor + lista stringova.

Cita PVSOL premium JSON export i generira:
  - formatiran .xlsx s listovima: Sazetak, PV polja, Invertori, Rezultati
    (+ opcionalno STRINGOVI list koji PanelStringsExcel.lsp cita izravno)
  - opcionalno zaseban .csv s popisom stringova za panelauto

Popis stringova: svaki fizicki inverter dobiva jedinstveni broj (IN1, IN2, ...);
inverteri s kolicinom > 1 se sekvencijalno renumeriraju. Oznaka stringa je
  IN{inverter}-M{mppt}-S{string}-{broj_panela}   npr. IN1-M1-S1-15

Pokretanje (GUI):       python pvsolToExcel.py
Pokretanje (headless):  python pvsolToExcel.py ulaz.json izlaz.xlsx
"""

import os
import re
import sys
import csv
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PVSOL izvozi valutu kao "£" (locale artefakt). Hrvatski projekti -> EUR.
CURRENCY = "EUR"

# ── Stilovi ─────────────────────────────────────────────────────────────────
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(name=FONT_NAME, bold=True, color="1F4E78", size=11)
BASE_FONT = Font(name=FONT_NAME, size=10)
TOTAL_FONT = Font(name=FONT_NAME, bold=True, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Stupci popisa stringova (CSV + STRINGOVI list). Prvi stupac = oznaka koju
# PanelStringsExcel.lsp parsira (car row).
STRING_COLUMNS = [
    ("Oznaka", "token"),
    ("Inverter", "inverter"),
    ("Model invertera", "model"),
    ("MPP", "mpp"),
    ("String", "string"),
    ("Broj panela", "panels"),
]


# ── Helperi ───────────────────────────────────────────────────────────────────
def g(d, *keys, default=None):
    """Sigurno gnijezdo .get() kroz niz kljuceva."""
    cur = d
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k)
        if cur is None:
            return default
    return cur


def unwrap(x):
    """PVSOL vrijednosti su {"Value": .., "Unit": ..}. Vraca (value, unit)."""
    if isinstance(x, dict):
        if "Value" in x:
            return x.get("Value"), loc_unit(x.get("Unit", ""))
        return "", ""
    return x, ""


def loc_unit(unit):
    """Lokalizira jedinicu: £ -> CURRENCY, Year -> god."""
    if not unit:
        return ""
    u = unit.replace("£", CURRENCY)
    u = re.sub(r"(?i)years?", "god", u)
    return u.strip()


def as_num(v):
    """Vraca int kad je cijeli broj, inace float; nenumericko ostaje kakvo je."""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return int(v) if float(v).is_integer() else float(v)
    return v


def num_fmt(v):
    return "#,##0" if isinstance(v, int) else "#,##0.00"


def fmt_date(s):
    if not s:
        return ""
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return s


def mpp_parse(cfg_list):
    """["MPP 1: 2 x 15", ...] -> {1: "2 x 15", ...}."""
    out = {}
    for item in cfg_list or []:
        if ":" not in item:
            continue
        left, right = item.split(":", 1)
        m = re.search(r"(\d+)", left)
        if m:
            out[int(m.group(1))] = right.strip()
    return out


def parse_n_x_m(s):
    """"2 x 15" -> (2, 15)  (broj stringova, broj panela po stringu)."""
    m = re.search(r"(\d+)\s*[x×]\s*(\d+)", str(s))
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r"(\d+)", str(s))
    return (1, int(m.group(1))) if m else (0, 0)


def put(ws, r, c, value, *, font=BASE_FONT, fill=None, fmt=None, align=None, border=False):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = Alignment(horizontal=align, vertical="center")
    if border:
        cell.border = BORDER
    return cell


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_headers(ws, headers, row=1):
    for c, h in enumerate(headers, start=1):
        put(ws, row, c, h, font=HEADER_FONT, fill=HEADER_FILL,
            align="center", border=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def write_kv_value(ws, r, value, unit):
    """Upisuje (vrijednost, jedinica) u stupce B i C key-value lista."""
    v = as_num(value)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        put(ws, r, 2, v, fmt=num_fmt(v), align="right", border=True)
    else:
        put(ws, r, 2, v if isinstance(v, str) else "", border=True)
    put(ws, r, 3, unit, border=True)


# ── Lista stringova ─────────────────────────────────────────────────────────
def build_string_rows(data):
    """Razvija PVSOL konfiguraciju u popis pojedinacnih stringova.

    Svaki fizicki inverter (uz uvazavanje kolicine) dobiva jedinstveni
    sekvencijalni broj. Vraca listu dict-ova sa kljucevima iz STRING_COLUMNS.
    """
    groups = g(data, "ProjectOverview", "Configuration", "ModuleAreas", default=[])
    rows = []
    phys = 0
    for grp in groups:
        inverters = sorted(grp.get("Inverters", []),
                           key=lambda x: x.get("Number", 0))
        for inv in inverters:
            qty = int(as_num(g(inv, "Quantity", "Value")) or 1)
            model = inv.get("Description", "")
            mpp_map = mpp_parse(inv.get("Configuration"))
            for _ in range(qty):
                phys += 1
                for mppt in sorted(mpp_map):
                    n_strings, panels = parse_n_x_m(mpp_map[mppt])
                    for s in range(1, n_strings + 1):
                        rows.append({
                            "token": f"IN{phys}-M{mppt}-S{s}-{panels}",
                            "inverter": phys,
                            "model": model,
                            "mpp": mppt,
                            "string": s,
                            "panels": panels,
                        })
    return rows


def write_strings_csv(rows, path):
    """Zapisuje popis stringova u CSV (utf-8-sig, zarez)."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([h for h, _ in STRING_COLUMNS])
        for r in rows:
            w.writerow([r[k] for _, k in STRING_COLUMNS])


def fill_stringovi(ws, rows):
    set_widths(ws, [18, 10, 30, 8, 8, 14])
    write_headers(ws, [h for h, _ in STRING_COLUMNS])
    for i, row in enumerate(rows, start=2):
        for c, (_, key) in enumerate(STRING_COLUMNS, start=1):
            val = row[key]
            align = "left" if key in ("token", "model") else "center"
            put(ws, i, c, val, align=align, border=True)


# ── Listovi ─────────────────────────────────────────────────────────────────
def fill_sazetak(ws, data):
    set_widths(ws, [34, 20, 16])
    write_headers(ws, ["Parametar", "Vrijednost", "Jedinica"])

    po = g(data, "ProjectOverview", default={})
    cons = g(po, "Consumption", default={})
    td = g(po, "ThreeDDesign", default={})
    cfg = g(po, "Configuration", default={})
    ov = g(data, "Results", "Overview", default={})
    econ = g(data, "Results", "FinancialAnalysis", "EconomicParameters", default={})
    pay = g(data, "Results", "FinancialAnalysis", "PaymentOverview", default={})

    rows = [
        ("PVSOL verzija", data.get("PVSOLVersion"), ""),
        ("Datum exporta", fmt_date(data.get("ExportDate")), ""),
        ("Lokacija / klima", g(po, "SystemTypeClimateAndGrid", "ClimateData"), ""),
        ("Tip sustava", g(po, "SystemTypeClimateAndGrid", "TypeOfSystem"), ""),
        ("Mreza", g(po, "SystemTypeClimateAndGrid", "AcMains"), ""),
        ("Pocetak rada", fmt_date(g(po, "ProjectData", "StartOfOperation", "Date")), ""),
        ("Ukupna snaga (DC)", *unwrap(g(td, "TotalPower", default={}))),
        ("Snaga (AC, konfiguracija)", *unwrap(g(cfg, "TotalPower", default={}))),
        ("Sizing faktor", *unwrap(g(cfg, "SizingFactor", default={}))),
        ("Godisnja potrosnja", *unwrap(g(cons, "TotalConsumption", default={}))),
        ("Vrsno opterecenje", *unwrap(g(cons, "LoadPeak", default={}))),
        ("Specificni prinos", *unwrap(g(ov, "SpecificAnnualYield", default={}))),
        ("Performance ratio", *unwrap(g(ov, "PerformanceRatio", default={}))),
        ("Proizvodnja PV (AC u mrezu)", *unwrap(g(ov, "PvGeneratorEnergyAcGrid", default={}))),
        ("Energija iz mreze", *unwrap(g(ov, "EnergyFromGrid", default={}))),
        ("Solarni udio", *unwrap(g(data, "Results", "Simulation", "Appliances", "SolarFraction", default={}))),
        ("Stupanj samodostatnosti", *unwrap(g(data, "Results", "Simulation", "LevelOfSelfSufficiency", "LevelOfSelfSufficiency", default={}))),
        ("CO2 izbjegnut", *unwrap(g(data, "Results", "Simulation", "PvSystem", "Co2EmissionsAvoided", default={}))),
        ("IRR (interna stopa povrata)", *unwrap(g(ov, "InternalRateOfReturn", default={}))),
        ("Godisnja usteda", *unwrap(g(ov, "RevenueOrSavings", default={}))),
        ("Period amortizacije", g(econ, "AmortizationPeriod", "Period"), ""),
        ("Investicija", *unwrap(g(pay, "InvestmentCosts", default={}))),
    ]
    for i, (name, value, unit) in enumerate(rows, start=2):
        put(ws, i, 1, name, border=True)
        write_kv_value(ws, i, value, unit)


def fill_pv_polja(ws, data):
    set_widths(ws, [44, 24, 14, 12, 12, 10, 14, 20])
    headers = ["Modulno polje", "Model panela", "Proizvodjac", "Broj panela",
               "Snaga (kWp)", "Nagib (st)", "Orijentacija (st)", "Tip montaze"]
    write_headers(ws, headers)

    areas = g(data, "ProjectOverview", "ThreeDDesign", "ModuleAreas", default=[])
    r = 2
    for a in areas:
        put(ws, r, 1, a.get("ModuleArea", ""), border=True)
        put(ws, r, 2, a.get("ModuleData", ""), border=True)
        put(ws, r, 3, a.get("Manufacturer", ""), border=True)
        put(ws, r, 4, as_num(g(a, "NumberOfPvModules", "Value")), fmt="#,##0", align="right", border=True)
        put(ws, r, 5, as_num(g(a, "PvGeneratorOutput", "Value")), fmt="#,##0.00", align="right", border=True)
        put(ws, r, 6, as_num(g(a, "Inclination", "Value")), fmt="0", align="right", border=True)
        put(ws, r, 7, as_num(g(a, "Orientation", "Value")), fmt="0", align="right", border=True)
        put(ws, r, 8, a.get("InstallationType", ""), border=True)
        r += 1

    if areas:
        put(ws, r, 1, "UKUPNO", font=TOTAL_FONT, border=True)
        for c in (2, 3, 6, 7, 8):
            put(ws, r, c, "", border=True)
        put(ws, r, 4, f"=SUM(D2:D{r - 1})", font=TOTAL_FONT, fmt="#,##0", align="right", border=True)
        put(ws, r, 5, f"=SUM(E2:E{r - 1})", font=TOTAL_FONT, fmt="#,##0.00", align="right", border=True)
    return len(areas)


def fill_invertori(ws, data):
    set_widths(ws, [6, 54, 30, 22, 10, 12, 12, 12, 12, 14])
    headers = ["Br.", "Modulno polje (grupa)", "Model invertera", "Proizvodjac",
               "Kolicina", "MPP 1", "MPP 2", "MPP 3", "MPP 4", "Sizing (%)"]
    write_headers(ws, headers)

    groups = g(data, "ProjectOverview", "Configuration", "ModuleAreas", default=[])
    r = 2
    count = 0
    for grp in groups:
        grp_name = grp.get("ModuleArea", "")
        for inv in grp.get("Inverters", []):
            put(ws, r, 1, as_num(inv.get("Number")), align="center", border=True)
            put(ws, r, 2, grp_name, border=True)
            put(ws, r, 3, inv.get("Description", ""), border=True)
            put(ws, r, 4, inv.get("Manufacturer", ""), border=True)
            put(ws, r, 5, as_num(g(inv, "Quantity", "Value")), fmt="#,##0", align="center", border=True)
            mpp = mpp_parse(inv.get("Configuration"))
            for k in range(1, 5):
                put(ws, r, 5 + k, mpp.get(k, ""), align="center", border=True)
            put(ws, r, 10, as_num(g(inv, "SizingFactor", "Value")), fmt="#,##0.0", align="right", border=True)
            r += 1
            count += 1
    return count


def fill_rezultati(ws, data):
    set_widths(ws, [40, 20, 16])
    write_headers(ws, ["Parametar", "Vrijednost", "Jedinica"])

    sim = g(data, "Results", "Simulation", default={})
    fin = g(data, "Results", "FinancialAnalysis", default={})
    pv = g(sim, "PvSystem", default={})
    app = g(sim, "Appliances", default={})
    sss = g(sim, "LevelOfSelfSufficiency", default={})
    sysd = g(fin, "SystemData", default={})
    rem = g(fin, "StartDurationAndEndOfRemuneration", default={})
    econ = g(fin, "EconomicParameters", default={})
    pay = g(fin, "PaymentOverview", default={})
    sav = g(fin, "RemunerationAndSavings", default={})

    tariff = (g(sav, "FromGridTariffs", default=[]) or [{}])[0]
    price = (g(tariff, "ElectricityPurchasePrices", default=[]) or [{}])[0]

    blocks = [
        ("PV sustav", [
            ("Snaga generatora", *unwrap(g(pv, "PvGeneratorOutput", default={}))),
            ("Specificni godisnji prinos", *unwrap(g(pv, "SpecificAnnualYield", default={}))),
            ("Performance ratio", *unwrap(g(pv, "PerformanceRatio", default={}))),
            ("Gubitak zbog sjene", *unwrap(g(pv, "YieldReductionDueToShading", default={}))),
            ("Proizvodnja u mrezu (s baterijom)", *unwrap(g(pv, "PvGeneratorEnergyAcGridWithBattery", default={}))),
            ("Clipping na tocki predaje", *unwrap(g(pv, "ClippingAtFeedInPoint", default={}))),
            ("CO2 izbjegnut", *unwrap(g(pv, "Co2EmissionsAvoided", default={}))),
        ]),
        ("Potrosnja", [
            ("Trosila", *unwrap(g(app, "Appliances", default={}))),
            ("Standby invertera", *unwrap(g(app, "StandbyConsumptionInverter", default={}))),
            ("Ukupna potrosnja", *unwrap(g(app, "TotalConsumption", default={}))),
            ("Energija iz mreze", *unwrap(g(app, "EnergyFromGrid", default={}))),
            ("Solarni udio", *unwrap(g(app, "SolarFraction", default={}))),
        ]),
        ("Samodostatnost", [
            ("Ukupna potrosnja", *unwrap(g(sss, "TotalConsumption", default={}))),
            ("Pokriveno iz mreze", *unwrap(g(sss, "CoveredByGrid", default={}))),
            ("Stupanj samodostatnosti", *unwrap(g(sss, "LevelOfSelfSufficiency", default={}))),
        ]),
        ("Financije", [
            ("Period procjene", *unwrap(g(sysd, "AssessmentPeriodInput", default={}))),
            ("Kamata na kapital", *unwrap(g(sysd, "InterestOnCapital", default={}))),
            ("Pocetak rada sustava", fmt_date(g(rem, "StartOfOperationOfTheSystem", "Date")), ""),
            ("Period remuneracije", g(rem, "RemunerationPeriod", "Period"), ""),
            ("Kraj remuneracije", fmt_date(g(rem, "EndOfRemuneration", "Date")), ""),
            ("IRR (interna stopa povrata)", *unwrap(g(econ, "InternalRateOfReturn", default={}))),
            ("Akumulirani novcani tok", *unwrap(g(econ, "AccruedCashFlow", default={}))),
            ("Period amortizacije", g(econ, "AmortizationPeriod", "Period"), ""),
            ("Trosak proizvodnje el. energije", *unwrap(g(econ, "ElectricityProductionCosts", default={}))),
            ("Specificna investicija", *unwrap(g(pay, "SpecificInvestmentCost", default={}))),
            ("Investicija", *unwrap(g(pay, "InvestmentCosts", default={}))),
            ("Usteda u 1. godini", *unwrap(g(sav, "FirstYearSavings", default={}))),
            ("Tarifa (iz mreze)", g(tariff, "FullName"), ""),
            ("Cijena energije", *unwrap(g(price, "EnergyPrice", default={}))),
            ("Inflacija cijene energije", *unwrap(g(tariff, "InflationRateForEnergyPrice", default={}))),
        ]),
    ]

    r = 2
    for title, items in blocks:
        for c in (1, 2, 3):
            put(ws, r, c, title if c == 1 else "", font=SECTION_FONT, fill=SECTION_FILL, border=True)
        r += 1
        for name, value, unit in items:
            put(ws, r, 1, name, border=True)
            write_kv_value(ws, r, value, unit)
            r += 1
    r += 1
    put(ws, r, 1, f"Napomena: PVSOL je izvezao valutu kao GBP; vrijednosti su prikazane kao {CURRENCY}.",
        font=Font(name=FONT_NAME, italic=True, size=9, color="808080"))


# ── Build ─────────────────────────────────────────────────────────────────────
def build_workbook(data, include_main=True, include_strings=True):
    """Gradi workbook. Vraca (wb, stats). string_rows u stats za daljnju upotrebu."""
    wb = Workbook()
    default = wb.active

    stats = {"areas": 0, "inverters": 0, "strings": 0, "phys_inverters": 0,
             "panels": 0, "string_rows": []}

    if include_main:
        fill_sazetak(wb.create_sheet("Sazetak"), data)
        stats["areas"] = fill_pv_polja(wb.create_sheet("PV polja"), data)
        stats["inverters"] = fill_invertori(wb.create_sheet("Invertori"), data)
        fill_rezultati(wb.create_sheet("Rezultati"), data)

    rows = build_string_rows(data)
    stats["string_rows"] = rows
    stats["strings"] = len(rows)
    stats["phys_inverters"] = max((r["inverter"] for r in rows), default=0)
    stats["panels"] = sum(r["panels"] for r in rows)

    if include_strings:
        fill_stringovi(wb.create_sheet("STRINGOVI"), rows)

    wb.remove(default)  # ukloni prazni default "Sheet"
    return wb, stats


# ── GUI (jednostavan, u stilu HEPA Form Fillera) ──────────────────────────────
class PvsolApp:
    def __init__(self, root):
        import tkinter as tk
        from tkinter import ttk
        self.tk, self.ttk = tk, ttk

        self.root = root
        self.root.title("PVSOL -> Excel")
        self.root.geometry("680x460")
        self.root.minsize(620, 420)

        self.json_path = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.make_xlsx = tk.BooleanVar(value=True)
        self.make_csv = tk.BooleanVar(value=True)
        self.make_stringovi = tk.BooleanVar(value=True)

        self._build_ui()

    def _build_ui(self):
        tk, ttk = self.tk, self.ttk

        # ── Red 1: odabir datoteka ───────────────────────────────────────
        row1 = ttk.Frame(self.root, relief="raised")
        row1.pack(fill="x")
        ttk.Label(row1, text="PVSOL JSON:", width=11).grid(row=0, column=0, padx=(8, 2), pady=5, sticky="w")
        ttk.Entry(row1, textvariable=self.json_path, width=52).grid(row=0, column=1, pady=5, sticky="we")
        ttk.Button(row1, text="…", width=3, command=self._pick_json).grid(row=0, column=2, padx=(2, 8))

        ttk.Label(row1, text="Spremi u:", width=11).grid(row=1, column=0, padx=(8, 2), pady=(0, 6), sticky="w")
        ttk.Entry(row1, textvariable=self.output_dir, width=52).grid(row=1, column=1, pady=(0, 6), sticky="we")
        ttk.Button(row1, text="…", width=3, command=self._pick_outdir).grid(row=1, column=2, padx=(2, 8))
        row1.columnconfigure(1, weight=1)

        # ── Red 2: opcije izlaza ─────────────────────────────────────────
        opts = ttk.LabelFrame(self.root, text="Sto generirati")
        opts.pack(fill="x", padx=8, pady=(8, 4))
        ttk.Checkbutton(opts, text="Excel tablica (Sazetak, PV polja, Invertori, Rezultati)",
                        variable=self.make_xlsx).pack(anchor="w", padx=8, pady=2)
        ttk.Checkbutton(opts, text="Lista stringova — CSV (za panelauto)",
                        variable=self.make_csv).pack(anchor="w", padx=8, pady=2)
        ttk.Checkbutton(opts, text="Lista stringova — list 'STRINGOVI' u xlsx (za PanelStringsExcel.lsp)",
                        variable=self.make_stringovi).pack(anchor="w", padx=8, pady=2)

        # ── Red 3: akcije ────────────────────────────────────────────────
        row3 = ttk.Frame(self.root)
        row3.pack(fill="x", padx=8, pady=4)
        ttk.Button(row3, text="\U0001F4C4  Generiraj", command=self._generate).pack(side="left")
        ttk.Button(row3, text="\U0001F4C2  Otvori mapu", command=self._open_outdir).pack(side="left", padx=6)

        # ── Sazetak (preview) ────────────────────────────────────────────
        sf = ttk.LabelFrame(self.root, text="Sadrzaj")
        sf.pack(fill="both", expand=True, padx=8, pady=4)
        self.tree = ttk.Treeview(sf, columns=("k", "v"), show="headings", height=8)
        self.tree.heading("k", text="Stavka")
        self.tree.heading("v", text="Vrijednost")
        self.tree.column("k", width=360, stretch=True)
        self.tree.column("v", width=200, stretch=True)
        self.tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(sf, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")

        # ── Status bar ───────────────────────────────────────────────────
        self.status = tk.StringVar(value="Odaberi PVSOL JSON datoteku.")
        ttk.Label(self.root, textvariable=self.status, relief="sunken", anchor="w").pack(
            fill="x", side="bottom", ipady=2)

    # ── callbacks ─────────────────────────────────────────────────────────
    def _pick_json(self):
        from tkinter import filedialog
        p = filedialog.askopenfilename(
            title="Odaberi PVSOL JSON export",
            filetypes=[("JSON datoteke", "*.json"), ("Sve datoteke", "*.*")])
        if p:
            self.json_path.set(p)
            if not self.output_dir.get():
                self.output_dir.set(os.path.dirname(p))
            self._preview()

    def _pick_outdir(self):
        from tkinter import filedialog
        p = filedialog.askdirectory(title="Izlazna mapa")
        if p:
            self.output_dir.set(p)

    def _open_outdir(self):
        d = self.output_dir.get()
        if d and os.path.isdir(d):
            os.startfile(d) if os.name == "nt" else None

    def _load_data(self):
        with open(self.json_path.get(), encoding="utf-8") as f:
            return json.load(f)

    def _fill_tree(self, items):
        self.tree.delete(*self.tree.get_children())
        for k, v in items:
            self.tree.insert("", "end", values=(k, v))

    def _preview(self):
        from tkinter import messagebox
        try:
            data = self._load_data()
        except Exception as e:
            messagebox.showerror("Greska", f"Ne mogu ucitati JSON:\n{e}")
            self.status.set("Greska pri citanju JSON-a.")
            return
        _, st = build_workbook(data, include_main=False, include_strings=False)
        proj = g(data, "ProjectOverview", "SystemTypeClimateAndGrid", "ClimateData") or "-"
        dc, _u = unwrap(g(data, "ProjectOverview", "ThreeDDesign", "TotalPower", default={}))
        self._fill_tree([
            ("Lokacija / klima", proj),
            ("Ukupna snaga (DC)", f"{dc} kWp"),
            ("PV polja (modulna)", st["areas"] or len(g(data, "ProjectOverview", "ThreeDDesign", "ModuleAreas", default=[]))),
            ("Invertori (PVSOL grupe)", st["inverters"] or sum(len(grp.get("Inverters", [])) for grp in g(data, "ProjectOverview", "Configuration", "ModuleAreas", default=[]))),
            ("Fizicki inverteri (IN1..)", st["phys_inverters"]),
            ("Stringovi (ukupno)", st["strings"]),
            ("Panela u stringovima", st["panels"]),
        ])
        self.status.set("Spreman za generiranje.")

    def _generate(self):
        from tkinter import messagebox
        if not self.json_path.get():
            messagebox.showwarning("Upozorenje", "Odaberi PVSOL JSON datoteku!")
            return
        if not self.output_dir.get():
            messagebox.showwarning("Upozorenje", "Odaberi izlaznu mapu!")
            return
        if not (self.make_xlsx.get() or self.make_csv.get() or self.make_stringovi.get()):
            messagebox.showwarning("Upozorenje", "Odaberi barem jedan izlaz!")
            return

        try:
            data = self._load_data()
        except Exception as e:
            messagebox.showerror("Greska", f"Ne mogu ucitati JSON:\n{e}")
            return

        base = os.path.splitext(os.path.basename(self.json_path.get()))[0]
        outdir = self.output_dir.get()
        made = []

        try:
            if self.make_xlsx.get() or self.make_stringovi.get():
                wb, st = build_workbook(data, include_main=self.make_xlsx.get(),
                                        include_strings=self.make_stringovi.get())
                xlsx_path = os.path.join(outdir, base + ".xlsx")
                wb.save(xlsx_path)
                made.append(xlsx_path)
            else:
                _, st = build_workbook(data, include_main=False, include_strings=False)

            if self.make_csv.get():
                csv_path = os.path.join(outdir, base + "_stringovi.csv")
                write_strings_csv(st["string_rows"], csv_path)
                made.append(csv_path)
        except Exception as e:
            messagebox.showerror("Greska", f"Greska pri generiranju:\n{e}")
            self.status.set("Greska pri generiranju.")
            return

        self._fill_tree([
            ("PV polja", st["areas"]),
            ("Invertori (PVSOL grupe)", st["inverters"]),
            ("Fizicki inverteri (IN1..)", st["phys_inverters"]),
            ("Stringovi (ukupno)", st["strings"]),
            ("Panela u stringovima", st["panels"]),
            *[("Datoteka", os.path.basename(p)) for p in made],
        ])
        self.status.set(f"Gotovo — generirano {len(made)} datoteka u {outdir}")
        if messagebox.askyesno("Gotovo", "Generirano:\n" + "\n".join(made) + "\n\nOtvoriti mapu?"):
            self._open_outdir()


def run_gui():
    import tkinter as tk
    root = tk.Tk()
    PvsolApp(root)
    root.mainloop()


def run_cli(in_path, out_path):
    with open(in_path, encoding="utf-8") as f:
        data = json.load(f)
    wb, st = build_workbook(data, include_main=True, include_strings=True)
    wb.save(out_path)
    csv_path = os.path.splitext(out_path)[0] + "_stringovi.csv"
    write_strings_csv(st["string_rows"], csv_path)
    print(f"Spremljeno: {out_path}")
    print(f"Spremljeno: {csv_path}")
    print(f"PV polja: {st['areas']} | invertori: {st['inverters']} | "
          f"fizicki: {st['phys_inverters']} | stringovi: {st['strings']} | "
          f"panela: {st['panels']}")


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        run_cli(sys.argv[1], sys.argv[2])
    else:
        run_gui()
