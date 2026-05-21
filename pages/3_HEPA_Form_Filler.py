"""
HEPA Form Filler – ispuni HEPA PDF obrasce iz Excel projektne dokumentacije.
"""

import io
import json
import streamlit as st
from openpyxl import load_workbook

try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

try:
    from PyPDFForm import PdfWrapper
    HAS_PYPDFFORM = True
except ImportError:
    HAS_PYPDFFORM = False

# ── Konstante ─────────────────────────────────────────────────────────────────

EXCEL_KEYWORDS = {
    "ime i prezime": "ime_prezime", "naziv tvrtke": "ime_prezime",
    "oib": "oib",
    "adresa priključnog": "adresa_pm", "adresa instalacije": "adresa_pm",
    "ulica i broj": "adresa_pm",
    "grad": "grad", "mjesto": "grad",
    "poštanski broj": "postanski_broj", "p.b.": "postanski_broj",
    "broj mjernog mjesta": "brm", "mmo": "brm", "brm": "brm",
    "instalirana snaga fne": "inst_snaga_fne", "ukupna snaga": "inst_snaga_fne",
    "zakupljena snaga": "zak_snaga",
    "broj panela": "br_panela",
    "snaga panela": "snaga_panela",
    "proizvođač panela": "pr_pan", "model panela": "md_pan",
    "broj invertera": "br_inv",
    "nazivna snaga invertera": "snaga_inv", "snaga invertera": "snaga_inv",
    "proizvođač invertera": "pr_inv", "model invertera": "md_inv",
    "mjesto i datum": "mj_dat",
    "telefon": "telefon", "mob": "telefon", "mobitel": "telefon",
    "e-mail": "email", "email": "email",
    "katastarska čestica": "kat_cestica", "k.č.": "kat_cestica",
    "katastarska općina": "kat_opcina", "k.o.": "kat_opcina",
}

TAG_LABELS = {
    "ime_prezime":    "Ime i prezime / Naziv tvrtke",
    "oib":            "OIB",
    "adresa_pm":      "Adresa priključnog mjesta",
    "grad":           "Grad / Mjesto",
    "postanski_broj": "Poštanski broj",
    "brm":            "Broj mjernog mjesta (MMO/BRM)",
    "inst_snaga_fne": "Instalirana snaga FNE (kW)",
    "zak_snaga":      "Zakupljena snaga (kW)",
    "br_panela":      "Broj panela",
    "snaga_panela":   "Snaga jednog panela (W)",
    "pr_pan":         "Proizvođač panela",
    "md_pan":         "Model panela",
    "br_inv":         "Broj invertera",
    "snaga_inv":      "Nazivna snaga invertera (kW)",
    "pr_inv":         "Proizvođač invertera",
    "md_inv":         "Model invertera",
    "mj_dat":         "Mjesto i datum",
    "telefon":        "Telefon / Mobitel",
    "email":          "E-mail",
    "kat_cestica":    "Katastarska čestica",
    "kat_opcina":     "Katastarska općina",
}

PDF_FIELD_HINTS = {
    "ime": "ime_prezime", "name": "ime_prezime", "naziv": "ime_prezime",
    "oib": "oib", "adresa": "adresa_pm", "address": "adresa_pm",
    "grad": "grad", "city": "grad", "mjesto": "grad",
    "postal": "postanski_broj", "postansk": "postanski_broj",
    "brm": "brm", "mjerno": "brm", "mmo": "brm",
    "inst": "inst_snaga_fne", "ukupna": "inst_snaga_fne",
    "zakup": "zak_snaga",
    "br_pan": "br_panela", "snaga_pan": "snaga_panela",
    "pr_pan": "pr_pan", "md_pan": "md_pan",
    "br_inv": "br_inv", "snaga_inv": "snaga_inv",
    "pr_inv": "pr_inv", "md_inv": "md_inv",
    "datum": "mj_dat", "tel": "telefon", "mob": "telefon",
    "email": "email", "mail": "email",
    "cestica": "kat_cestica", "opcina": "kat_opcina",
}

# ── Logika ────────────────────────────────────────────────────────────────────

def extract_from_excel(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = next((n for n in wb.sheetnames if "podaci" in n.lower()), None)
    if sheet is None:
        return {}
    ws = wb[sheet]
    found = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            val = cell.value
            if not isinstance(val, str):
                continue
            t = val.strip().lower()
            for kw, tag in EXCEL_KEYWORDS.items():
                if kw in t and tag not in found:
                    nxt = ws.cell(row=cell.row, column=cell.column + 1).value
                    if nxt is not None and str(nxt).strip():
                        found[tag] = str(nxt).strip()
    wb.close()
    return found


def get_form_fields(pdf_bytes):
    if not HAS_PYPDFFORM:
        return {}
    try:
        w = PdfWrapper(io.BytesIO(pdf_bytes))
        return dict(w.data) if w.data else {}
    except Exception:
        return {}


def auto_map_form_fields(form_fields, excel_data):
    out = {}
    for field in form_fields:
        fl = field.lower().replace("-", "_").replace(" ", "_")
        for hint, tag in PDF_FIELD_HINTS.items():
            if hint in fl and tag in excel_data:
                out[field] = excel_data[tag]
                break
    return out


def generate_fillable(pdf_bytes, field_values):
    w = PdfWrapper(io.BytesIO(pdf_bytes))
    valid = {k: v for k, v in field_values.items() if k in (w.data or {})}
    out = w.fill(valid)
    return out.read()


def generate_overlay(pdf_bytes, layout, excel_data):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    for tag, cfg in layout.get("fields", {}).items():
        page_i = int(cfg.get("page", 0))
        if page_i >= len(doc):
            continue
        value = excel_data.get(tag) or cfg.get("manual_value", "")
        if not value:
            continue
        page = doc[page_i]
        page.insert_text(
            fitz.Point(float(cfg["x"]), float(cfg["y"])),
            str(value),
            fontsize=int(cfg.get("font_size", 10)),
            color=(0, 0, 0),
        )
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def render_pdf_page(pdf_bytes, page_idx=0, dpi=120, layout=None, excel_data=None):
    """Renderira PDF stranicu u PNG bytes, opcionalno s overlay poljem."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc[page_idx]
    if layout and excel_data is not None:
        for tag, cfg in layout.get("fields", {}).items():
            if int(cfg.get("page", 0)) != page_idx:
                continue
            val = excel_data.get(tag) or cfg.get("manual_value", "") or f"[{TAG_LABELS.get(tag, tag)}]"
            page.insert_text(
                fitz.Point(float(cfg["x"]), float(cfg["y"])),
                str(val),
                fontsize=int(cfg.get("font_size", 10)),
                color=(0.8, 0, 0),
            )
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page.get_pixmap(matrix=mat, alpha=False)
    return pix.tobytes("png"), page.rect.width, page.rect.height


# ── UI ────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title='HEPA Form Filler', page_icon='📋', layout='wide')
st.title('📋 HEPA Form Filler')
st.caption('Automatski ispunjava HEPA PDF obrasce iz Excel projektne dokumentacije')
st.divider()

if not HAS_FITZ:
    st.error('Nedostaje PyMuPDF. Pokreni: pip install PyMuPDF')
    st.stop()

# ── Upload ────────────────────────────────────────────────────────────────────
col_xl, col_pdf = st.columns(2)
with col_xl:
    excel_file = st.file_uploader('Excel projektna datoteka (.xlsx/.xlsm)', type=['xlsx', 'xlsm'])
with col_pdf:
    pdf_file = st.file_uploader('HEPA PDF predložak', type=['pdf'])

if not excel_file or not pdf_file:
    st.info('Učitaj i Excel i PDF datoteku za nastavak.')
    st.stop()

# ── Učitaj podatke ────────────────────────────────────────────────────────────
excel_bytes = excel_file.read()
pdf_bytes   = pdf_file.read()

with st.spinner('Čitam Excel...'):
    try:
        excel_data = extract_from_excel(excel_bytes)
    except Exception as e:
        st.error(f'Greška pri čitanju Excela: {e}')
        st.stop()

form_fields = get_form_fields(pdf_bytes)
pdf_type    = 'fillable' if form_fields else 'flat'

st.success(f'Učitano **{len(excel_data)}** Excel polja  |  PDF tip: **{pdf_type}**')

# ── Prikaz Excel podataka ─────────────────────────────────────────────────────
with st.expander('Pregledaj učitane Excel podatke', expanded=True):
    import pandas as pd
    if excel_data:
        df = pd.DataFrame([
            {'Polje': TAG_LABELS.get(tag, tag), 'Vrijednost': val}
            for tag, val in excel_data.items()
        ])
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        st.warning('Nema pronađenih podataka. Provjeri ima li Excel list s imenom "Podaci".')

# ── Fillable PDF: direktno generiraj ──────────────────────────────────────────
if pdf_type == 'fillable' and HAS_PYPDFFORM:
    st.subheader('Fillable PDF – automatsko popunjavanje')
    mapped = auto_map_form_fields(form_fields, excel_data)

    with st.expander('Pregled mapiranih form polja'):
        if mapped:
            st.dataframe(pd.DataFrame([
                {'Form polje': k, 'Vrijednost': v} for k, v in mapped.items()
            ]), use_container_width=True, hide_index=True)
        else:
            st.warning('Nijedno form polje nije automatski mapirano.')

    if st.button('Generiraj ispunjeni PDF', type='primary'):
        with st.spinner('Generiram PDF...'):
            try:
                out_bytes = generate_fillable(pdf_bytes, mapped)
                st.download_button(
                    '⬇ Preuzmi ispunjeni PDF',
                    data=out_bytes,
                    file_name=f'HEPA_ispunjen_{excel_file.name}.pdf',
                    mime='application/pdf',
                )
            except Exception as e:
                st.error(f'Greška: {e}')

# ── Flat PDF: overlay editor ───────────────────────────────────────────────────
else:
    st.subheader('Overlay editor – postavljanje polja na PDF')
    st.caption('Postavi X/Y koordinate za svako polje. Koordinate su u PDF točkama (pt). Pregled se ažurira automatski.')

    # Layout session state
    if 'hepa_layout' not in st.session_state or st.session_state.get('hepa_pdf_name') != pdf_file.name:
        st.session_state.hepa_layout = {'fields': {}}
        st.session_state.hepa_pdf_name = pdf_file.name

    layout = st.session_state.hepa_layout

    # Broj stranica
    doc_tmp = fitz.open(stream=pdf_bytes, filetype='pdf')
    n_pages = len(doc_tmp)
    doc_tmp.close()

    col_ctrl, col_prev = st.columns([1, 2])

    with col_ctrl:
        page_idx = st.number_input('Stranica PDF-a', min_value=0, max_value=n_pages - 1, value=0, step=1)

        st.markdown('**Dodaj polje:**')
        all_tags = sorted(TAG_LABELS.keys(), key=lambda t: TAG_LABELS[t])
        sel_tag  = st.selectbox('Odaberi polje', all_tags, format_func=lambda t: TAG_LABELS[t])

        fc1, fc2, fc3 = st.columns(3)
        new_x  = fc1.number_input('X (pt)', value=100.0, step=5.0)
        new_y  = fc2.number_input('Y (pt)', value=100.0, step=5.0)
        new_fs = fc3.number_input('Font', value=10, min_value=5, max_value=48)

        manual_val = st.text_input('Ručna vrijednost (ako nije u Excelu)', value='')

        if st.button('➕ Dodaj polje na stranicu'):
            layout['fields'][sel_tag] = {
                'x': new_x, 'y': new_y,
                'font_size': new_fs,
                'page': page_idx,
                'manual_value': manual_val,
            }
            st.session_state.hepa_layout = layout
            st.rerun()

        st.divider()

        # Popis postavljenih polja
        fields_on_page = {t: c for t, c in layout['fields'].items() if int(c.get('page', 0)) == page_idx}
        if fields_on_page:
            st.markdown(f'**Polja na stranici {page_idx}:**')
            for tag, cfg in list(fields_on_page.items()):
                val = excel_data.get(tag) or cfg.get('manual_value', '') or '—'
                r1, r2 = st.columns([3, 1])
                r1.markdown(f'`{TAG_LABELS.get(tag, tag)}`  \n{val}')
                if r2.button('🗑', key=f'del_{tag}'):
                    del layout['fields'][tag]
                    st.session_state.hepa_layout = layout
                    st.rerun()
        else:
            st.info('Nema polja na ovoj stranici.')

        # Layout JSON import/export
        st.divider()
        layout_json = json.dumps(layout, ensure_ascii=False, indent=2)
        st.download_button('⬇ Spremi raspored (JSON)', data=layout_json,
                           file_name='hepa_layout.json', mime='application/json')
        uploaded_layout = st.file_uploader('Učitaj raspored (JSON)', type='json', key='layout_upload')
        if uploaded_layout:
            st.session_state.hepa_layout = json.load(uploaded_layout)
            st.rerun()

    with col_prev:
        st.markdown('**Pregled stranice:**')
        try:
            img_bytes, pw, ph = render_pdf_page(
                pdf_bytes, page_idx,
                layout=layout, excel_data=excel_data
            )
            st.image(img_bytes, use_container_width=True,
                     caption=f'Stranica {page_idx + 1} – dimenzije: {pw:.0f}×{ph:.0f} pt')
        except Exception as e:
            st.error(f'Greška pri renderiranju: {e}')

    st.divider()

    # Generiraj PDF
    if st.button('📄 Generiraj ispunjeni PDF', type='primary'):
        if not layout.get('fields'):
            st.warning('Dodaj barem jedno polje prije generiranja.')
        else:
            with st.spinner('Generiram PDF...'):
                try:
                    out_bytes = generate_overlay(pdf_bytes, layout, excel_data)
                    st.success('PDF generiran!')
                    st.download_button(
                        '⬇ Preuzmi ispunjeni PDF',
                        data=out_bytes,
                        file_name=f'HEPA_ispunjen_{excel_file.name}.pdf',
                        mime='application/pdf',
                    )
                except Exception as e:
                    st.error(f'Greška: {e}')
