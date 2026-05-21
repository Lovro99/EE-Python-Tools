import os, glob, csv
from pathlib import Path
from openpyxl import load_workbook

# GUI dijalozi
import tkinter as tk
from tkinter import filedialog

TARGETS = {
    "proizvođač invertera": "pr_inv",
    "model invertera": "md_inv",
    "proizvođač panela": "pr_pan",
    "model panela": "md_pan",
    "zakupljena snaga": "zak_snaga",
    "instalirana snaga fne": "inst_snaga_fne",  # novi ključ
}


def extract_from_file(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if "Podaci" not in wb.sheetnames:
            return None
        ws = wb["Podaci"]
        found = {v: None for v in TARGETS.values()}

        for row in ws.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str):
                    t = val.strip().lower()
                    for key, tag in TARGETS.items():
                        if key in t and found[tag] is None:
                            found[tag] = ws.cell(row=cell.row, column=cell.column + 1).value
            if all(found[k] is not None for k in found):
                break

        return [
            found["pr_inv"],
            found["md_inv"],
            found["pr_pan"],
            found["md_pan"],
            found["zak_snaga"],
            found["inst_snaga_fne"],  # ubaceno prije path
            os.path.abspath(path),
        ]
    finally:
        wb.close()

def run_recursive(root, out_csv):
    rows = []
    for path in glob.glob(str(Path(root) / "**" / "*.xlsm"), recursive=True):
        if Path(path).name.startswith("~"):
            continue
        rec = extract_from_file(path)
        if rec is not None:
            rows.append(rec)

    # zapis CSV
    with open(out_csv, "w", newline="", encoding="utf-16") as f:
        w = csv.writer(f)
        w.writerow([
            "Proizvođač Invertera",
            "Model Invertera",
            "Proizvođač Panela",
            "Model Panela",
            "Zakupljena snaga",
            "Instalirana snaga FNE",  # novo zaglavlje
            "Path"
        ])
        w.writerows(rows)

if __name__ == "__main__":
    # inicijaliziraj tkinter bez glavnog prozora
    root_tk = tk.Tk()
    root_tk.withdraw()

    # odaberi root mapu
    selected_root = filedialog.askdirectory(title="Odaberite root mapu za pretraživanje (.xlsm)")
    if not selected_root:
        print("Nije odabrana mapa. Prekid.")
        raise SystemExit(0)

    # predloži zadano ime CSV-a u toj mapi
    default_csv = os.path.join(selected_root, "rezultat.csv")

    # odaberi izlazni CSV
    out_csv_path = filedialog.asksaveasfilename(
        title="Spremi rezultat kao CSV",
        defaultextension=".csv",
        initialfile="rezultat.csv",
        initialdir=selected_root,
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
    )
    if not out_csv_path:
        print("Nije odabran izlazni CSV. Prekid.")
        raise SystemExit(0)

    run_recursive(selected_root, out_csv_path)
    print(f"Gotovo. Rezultat: {out_csv_path}")
