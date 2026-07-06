"""Ekstrakcija iz key-value lista (npr. Excel 'Podaci': A=naziv, B=vrijednost).

Za projekte gdje podaci nisu tablica krugova nego popis svojstava
projekta (FNE: proizvodac panela, model invertera, snage, kabeli...).
Spajanje je po KANONSKOM imenu polja iz zajednickog rjecnika (fields.py),
a pseudo-oznaka je 'PROJEKT'.

Primjer predloska:
    ime: podaci_kv
    tip: excel_kv
    datoteka: "*podat*.xls*"
    list: "Podaci"
    kolona_naziv: A
    kolona_vrijednost: B
    # koristi zajednicki 'polja:' rjecnik iz configa
"""

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from ..normalize import normalize_tag, normalize_value

TAG = "PROJEKT"


def extract_excel_kv(path, template, field_dict=None):
    if not field_dict:
        raise ValueError(
            f"predlozak '{template['ime']}' (excel_kv) zahtijeva "
            f"'polja:' rjecnik u configu"
        )
    name_col = column_index_from_string(template.get("kolona_naziv", "A"))
    val_col = column_index_from_string(template.get("kolona_vrijednost", "B"))
    sheet = template.get("list")

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet and sheet not in wb.sheetnames:
            raise ValueError(
                f"{path}: list '{sheet}' ne postoji "
                f"(dostupni: {', '.join(wb.sheetnames)})"
            )
        ws = wb[sheet] if sheet else wb.active
        sheet = ws.title

        rows = []
        seen_fields = set()
        max_col = max(name_col, val_col)
        # values_only izbjegava EmptyCell (nema .column) u read_only modu
        for r_idx, values in enumerate(
            ws.iter_rows(min_row=1, min_col=1, max_col=max_col, values_only=True),
            start=1,
        ):
            name_val = values[name_col - 1] if len(values) >= name_col else None
            val = values[val_col - 1] if len(values) >= val_col else None
            if name_val is None or val is None:
                continue
            field = field_dict.match(name_val)
            if field is None:
                continue
            raw_value = str(val).strip()
            if not raw_value:
                continue
            if field in seen_fields:  # prvo pojavljivanje ima prednost
                continue
            seen_fields.add(field)
            coord = f"{get_column_letter(val_col)}{r_idx}"
            rows.append({
                "tag": TAG,
                "raw_tag": TAG,
                "attribute": field,
                "value": normalize_value(
                    field, raw_value, numeric=field_dict.is_numeric(field)
                ),
                "raw_value": raw_value,
                "location": f"{sheet}!{coord} ({str(name_val).strip()})",
            })
        return rows
    finally:
        wb.close()
