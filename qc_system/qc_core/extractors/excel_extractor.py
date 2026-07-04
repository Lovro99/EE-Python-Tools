"""Ekstrakcija iz Excel bilanci/tablica prema opisu predloska.

Predlozak (qc_config.yaml) kaze koji list, od kojeg reda i koja kolona
znaci koji atribut. Kolona s vrijednoscu "oznaka" je kljuc; citanje
staje kad je oznaka prazna.

Primjer predloska:
    ime: bilanca_snaga
    tip: excel
    datoteka: "*bilanca*.xlsx"
    list: "Bilanca"          # ili izostavi za aktivni list
    prvi_red: 3
    kolone:
      A: oznaka
      C: snaga_kw
      E: struja_a
      F: tip_kabela
"""

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from ..normalize import normalize_tag, normalize_value


def extract_excel(path, template):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet_name = template.get("list")
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"{path}: list '{sheet_name}' ne postoji "
                    f"(dostupni: {', '.join(wb.sheetnames)})"
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        kolone = template["kolone"]
        col_map = {column_index_from_string(k): v for k, v in kolone.items()}
        key_col = next(
            (idx for idx, attr in col_map.items() if attr == "oznaka"), None
        )
        if key_col is None:
            raise ValueError(f"predlozak '{template['ime']}': nema kolone 'oznaka'")

        first_row = int(template.get("prvi_red", 2))
        rows = []
        for row in ws.iter_rows(min_row=first_row):
            cells = {c.column: c for c in row if c.column in col_map}
            key_cell = cells.get(key_col)
            raw_tag = "" if key_cell is None or key_cell.value is None \
                else str(key_cell.value).strip()
            if not raw_tag:
                break  # kraj tablice
            tag = normalize_tag(raw_tag)
            for col_idx, attribute in col_map.items():
                if attribute == "oznaka":
                    continue
                cell = cells.get(col_idx)
                if cell is None or cell.value is None:
                    continue
                raw_value = str(cell.value).strip()
                if not raw_value:
                    continue
                rows.append({
                    "tag": tag,
                    "raw_tag": raw_tag,
                    "attribute": attribute,
                    "value": normalize_value(attribute, raw_value),
                    "raw_value": raw_value,
                    "location": f"{sheet_name}!{cell.coordinate}",
                })
        return rows
    finally:
        wb.close()
