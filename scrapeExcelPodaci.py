import os
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import warnings

warnings.filterwarnings('ignore')

# Ključne riječi koje trebam pronaći
#KEYWORDS = [
#    'Proizvođač panela',
#    'Model panela',
#    'Snaga panela',
#    'Težina panela',
#    'Dimenzije panela',
#    'Broj ćelija panela',
#    'Nazivni napon-Umpp panela',
#    'Nazivna struja-Impp panela',
  #  'Napon otvorenog kruga-Uoc panela',
 #   'Struja kratkog spoja-Isc panela',
#    'Tip ćelije',
 #   'Efikasnost panela'
#]

KEYWORDS = [
    'Proizvođač invertera',
    'Broj invertera',
    'Model invertera',
    'Nazivna snaga invertera',
    'Raspon napona MPPT',
    'Radna temperatura',
    'Vlažnost zraka',
    'Prekidač invertera',
    'RDC invertera',
    'Izlazna struja invertera',
    'Maksimalna izlazna struja',
    'Kabel od invertera 1 do RO-SE'
]


def odaberi_mapu():
    """Otvara prozor za odabir mape"""
    root = tk.Tk()
    root.withdraw()
    mapa = filedialog.askdirectory(title="Odaberite mapu za pretragu .xlsm datoteka")
    root.destroy()
    return mapa

def odaberi_lokaciju_i_ime():
    """Otvara prozor za odabir lokacije i imena CSV datoteke"""
    root = tk.Tk()
    root.withdraw()
    
    lokacija = filedialog.askdirectory(title="Odaberite gdje spremiti CSV datoteku")
    
    if not lokacija:
        root.destroy()
        return None, None
    
    # Prozor za unos imena datoteke
    top = tk.Toplevel(root)
    top.title("Unos imena datoteke")
    top.geometry("400x150")
    top.attributes('-topmost', True)
    
    tk.Label(top, text="Kako trebam nazvati CSV datoteku?", font=("Arial", 10)).pack(pady=10)
    tk.Label(top, text="(bez .csv ekstenzije)", font=("Arial", 8), fg="gray").pack()
    
    entry = tk.Entry(top, width=40, font=("Arial", 10))
    entry.pack(pady=10)
    entry.insert(0, "panel_podaci")
    
    rezultat = []
    
    def spremi():
        naziv = entry.get().strip()
        if naziv:
            rezultat.append(naziv)
            top.destroy()
        else:
            messagebox.showwarning("Greška", "Molimo unesite naziv datoteke!")
    
    tk.Button(top, text="OK", command=spremi, bg="#4CAF50", fg="white", width=20).pack(pady=10)
    
    root.wait_window(top)
    root.destroy()
    
    return lokacija, rezultat[0] if rezultat else None

def pronadiKeyword(ws, keyword):
    """Pronalazi ključnu riječ u worksheetu i vraća vrijednost iz sljedećeg stupca"""
    try:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and str(cell.value).strip().lower() == keyword.lower():
                    # Pronađen keyword, dohvati vrijednost iz sljedećeg stupca
                    next_col = cell.column + 1
                    if next_col <= ws.max_column:
                        vrijednost = ws.cell(row=cell.row, column=next_col).value
                        return vrijednost
    except Exception as e:
        print(f"Greška pri pretrazi: {e}")
    
    return None

def proces_excel_datoteke(putanja_datoteke):
    """Obrađuje jednu Excel datoteku i vraća rječnik s podatcima"""
    try:
        wb = load_workbook(putanja_datoteke, read_only=False, data_only=True)
        
        # Provjeri postoji li sheet 'Podaci'
        if 'Podaci' not in wb.sheetnames:
            print(f"⚠️  UPOZORENJE: Sheet 'Podaci' ne postoji u {putanja_datoteke}")
            return None
        
        ws = wb['Podaci']
        
        # Spremi podatke
        red_podataka = {'Datoteka': os.path.basename(putanja_datoteke)}
        
        # Za svaki keyword pronađi vrijednost
        for keyword in KEYWORDS:
            vrijednost = pronadiKeyword(ws, keyword)
            red_podataka[keyword] = vrijednost if vrijednost is not None else ""
        
        wb.close()
        return red_podataka
    
    except PermissionError:
        print(f"⚠️  UPOZORENJE: Datoteka je zaključana ili nemaš pristup: {putanja_datoteke}")
        return None
    except Exception as e:
        print(f"⚠️  UPOZORENJE: Greška pri obradi {putanja_datoteke}: {e}")
        return None

def pretragi_xlsm_datoteke(pocetna_mapa):
    """Rekurzivno pretražuje mapu i traži sve .xlsm datoteke"""
    xlsm_datoteke = []
    
    print(f"\n🔍 Pretraživanje od: {pocetna_mapa}\n")
    
    for korijen, direktoriji, datoteke in os.walk(pocetna_mapa):
        for datoteka in datoteke:
            if datoteka.lower().endswith('.xlsm'):
                puna_putanja = os.path.join(korijen, datoteka)
                xlsm_datoteke.append(puna_putanja)
                print(f"✓ Pronađena: {puna_putanja}")
    
    return xlsm_datoteke

def main():
    print("=" * 60)
    print("EKSTRATOR PODATAKA IZ EXCEL DATOTEKA - PANELI")
    print("=" * 60)
    
    # Odabir polazne mape
    pocetna_mapa = odaberi_mapu()
    if not pocetna_mapa:
        print("❌ Pretraživanje otkazano - mapa nije odabrana!")
        return
    
    # Odabir lokacije i imena CSV datoteke
    lokacija_csv, naziv_csv = odaberi_lokaciju_i_ime()
    if not lokacija_csv or not naziv_csv:
        print("❌ Spremanje otkazano - lokacija nije odabrana!")
        return
    
    # Rekurzivno pretraži .xlsm datoteke
    xlsm_datoteke = pretragi_xlsm_datoteke(pocetna_mapa)
    
    if not xlsm_datoteke:
        print("❌ Nisu pronađene .xlsm datoteke!")
        return
    
    ukupno_datoteka = len(xlsm_datoteke)
    print(f"\n📊 Pronađeno {ukupno_datoteka} .xlsm datoteka(e)")
    print("\n⚙️  Obrada datoteka...\n")
    
    # Obrada svake datoteke
    svi_podaci = []
    za_uspjesno = 0
    
    for indeks, datoteka in enumerate(xlsm_datoteke, 1):
        # Prikazi praćenje napretka
        procenat = (indeks / ukupno_datoteka) * 100
        print(f"[{indeks}/{ukupno_datoteka}] ({procenat:.1f}%) Obrada: {os.path.basename(datoteka)}", end=" ")
        
        podatci = proces_excel_datoteke(datoteka)
        if podatci:
            svi_podaci.append(podatci)
            za_uspjesno += 1
            print("✓")
        else:
            print("✗")
    
    # Kreiraj DataFrame i spremi kao CSV
    if svi_podaci:
        df = pd.DataFrame(svi_podaci)
        
        # Postavi redoslijed stupaca
        stupci = ['Datoteka'] + KEYWORDS
        df = df[stupci]
        
        # Spremi CSV s UTF-8-sig encoding
        putanja_csv = os.path.join(lokacija_csv, f"{naziv_csv}.csv")
        df.to_csv(putanja_csv, index=False, encoding='utf-8-sig')
        
        print(f"\n{'=' * 60}")
        print(f"✅ USPJEŠNO!")
        print(f"{'=' * 60}")
        print(f"📁 Lokacija: {putanja_csv}")
        print(f"📈 Pronađenih datoteka: {ukupno_datoteka}")
        print(f"✓ Obrađenih datoteka: {za_uspjesno}/{ukupno_datoteka}")
        print(f"✗ Preskoči/Greške: {ukupno_datoteka - za_uspjesno}/{ukupno_datoteka}")
        print(f"📝 Redaka u CSV: {len(df)}")
        print(f"📊 Stupaca: {len(df.columns)}")
        print(f"{'=' * 60}")
        
        messagebox.showinfo("Uspješno!", 
                          f"CSV datoteka je uspješno kreirana!\n\n"
                          f"Lokacija: {putanja_csv}\n"
                          f"Pronađenih datoteka: {ukupno_datoteka}\n"
                          f"Obrađenih datoteka: {za_uspjesno}/{ukupno_datoteka}\n"
                          f"Redaka: {len(df)}")
    else:
        print("❌ Nisu obrađene nikakve datoteke!")
        messagebox.showerror("Greška", "Nisu obrađene nikakve datoteke!")

if __name__ == "__main__":
    main()