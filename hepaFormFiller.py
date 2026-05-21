"""
HEPA Form Filler v4
Automatski ispunjava HEPA PDF obrasce iz Excel projektne dokumentacije.

Ovisnosti:  pip install PyMuPDF Pillow openpyxl PyPDFForm
Pokretanje: python hepaFormFiller.py
"""

import os
import re
import json
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
import fitz
from PIL import Image, ImageTk

try:
    from PyPDFForm import PdfWrapper
    HAS_PYPDFFORM = True
except ImportError:
    HAS_PYPDFFORM = False

# ---------------------------------------------------------------------------
# Konstante
# ---------------------------------------------------------------------------

LAYOUT_SUFFIX = "_layout.json"
MIN_SCALE     = 0.3
MAX_SCALE     = 3.0

# TrueType fontovi s Unicode podrškom (za dijakritike u PDF outputu)
FONT_CANDIDATES = [
    r"C:\Windows\Fonts\arial.ttf",
    r"C:\Windows\Fonts\calibri.ttf",
    r"C:\Windows\Fonts\times.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
    "/usr/share/fonts/truetype/ubuntu/Ubuntu-R.ttf",
    "/System/Library/Fonts/Supplemental/Arial.ttf",
    "/Library/Fonts/Arial.ttf",
]

def find_unicode_font() -> str | None:
    for f in FONT_CANDIDATES:
        if os.path.exists(f):
            return f
    return None

# Globalna putanja do fonta — korisnik može promijeniti u GUI
_font_file: str | None = find_unicode_font()

# Excel ključne riječi → interni tagovi
EXCEL_KEYWORDS = {
    "ime investitora":               "ime_prezime",
    "naziv investitora":             "ime_prezime",
    "ime i prezime":                 "ime_prezime",
    "naziv tvrtke":                  "ime_prezime",
    "adresa investitora (ulica)":    "adresa_ulica",
    "ulica i broj":                  "adresa_ulica",
    "adresa investitora (grad)":     "adresa_grad",
    "oib investitora":               "oib",
    "oib":                           "oib",
    "adresa građevine":              "adresa_pm",
    "adresa priključnog":            "adresa_pm",
    "lokacija građevine":            "lokacija",
    "broj omm":                      "brm",
    "broj mjernog mjesta":           "brm",
    "mmo":                           "brm",
    "zakupljena snaga":              "zak_snaga",
    "vrsta priključka":              "vrsta_priklj",
    "instalirana snaga fne":         "inst_snaga_fne",
    "dc snaga elektrane":            "dc_snaga",
    "ac snaga elektrane":            "ac_snaga",
    "broj panela":                   "br_panela",
    "snaga panela":                  "snaga_panela",
    "proizvođač panela":             "pr_pan",
    "model panela":                  "md_pan",
    "tip ćelije":                    "tip_celije",
    "orijentacija panela":           "orijentacija",
    "kut nagiba":                    "kut_nagiba",
    "broj invertera":                "br_inv",
    "nazivna snaga invertera":       "snaga_inv",
    "snaga invertera":               "snaga_inv",
    "proizvođač invertera":          "pr_inv",
    "model invertera":               "md_inv",
    "mjesto i datum":                "mj_dat",
    "telefon":                       "telefon",
    "mobitel":                       "telefon",
    "mob":                           "telefon",
    "e-mail":                        "email",
    "email":                         "email",
    "glavni projektant":             "projektant",
    "projektant":                    "projektant",
    "predaja u el. mrežu":           "predana_energija",
    "predaja u mrežu":               "predana_energija",
    "preuzeta energija iz mreže":    "preuzeta_energija",
    "preuzeta energija":             "preuzeta_energija",
    "razlika između potrošnje i proizvodnje": "razlika_energija",
    "razlika između potrošnje":      "razlika_energija",
}

TAG_LABELS = {
    # Investitor (direktno iz Excela)
    "ime_prezime":   "Investitor – Ime / Naziv tvrtke",
    "oib":           "Investitor – OIB",
    "adresa_ulica":  "Investitor – Adresa ulica+broj (kombinirana)",
    "adresa_grad":   "Investitor – Grad+poštanski (kombiniran)",
    # Investitor (split)
    "inv_ulica":     "Investitor – Ulica",
    "inv_broj":      "Investitor – Kućni broj",
    "inv_post":      "Investitor – Poštanski broj",
    "inv_grad":      "Investitor – Grad / Mjesto",
    # Kontakt
    "telefon":       "Telefon / Mobitel",
    "email":         "E-mail",
    # Projektant
    "projektant":    "Projektant / Opunomoćenik",
    # Građevina (direktno)
    "adresa_pm":     "Građevina – Adresa (kombinirana)",
    "lokacija":      "Građevina – Lokacija (k.č./k.o. kombinirana)",
    # Građevina (split)
    "pm_ulica":      "Građevina – Ulica",
    "pm_broj":       "Građevina – Kućni broj",
    "pm_post":       "Građevina – Poštanski broj",
    "pm_grad":       "Građevina – Grad / Mjesto",
    "kat_cestica":   "Katastarska čestica",
    "kat_opcina":    "Katastarska općina",
    # Priključak
    "brm":           "Broj OMM / mjernog mjesta",
    "zak_snaga":     "Zakupljena snaga (kW)",
    "vrsta_priklj":  "Vrsta priključka",
    # FNE
    "inst_snaga_fne":"Instalirana snaga FNE",
    "dc_snaga":      "DC snaga elektrane (kWp)",
    "ac_snaga":      "AC snaga elektrane (kW)",
    "br_panela":     "Broj panela",
    "snaga_panela":  "Snaga jednog panela (W)",
    "pr_pan":        "Proizvođač panela",
    "md_pan":        "Model panela",
    "tip_celije":    "Tip ćelije",
    "orijentacija":  "Orijentacija panela",
    "kut_nagiba":    "Kut nagiba (°)",
    "br_inv":        "Broj invertera",
    "snaga_inv":     "Nazivna snaga invertera",
    "pr_inv":        "Proizvođač invertera",
    "md_inv":        "Model invertera",
    # Opće
    "mj_dat":        "Mjesto i datum",
    # Energija
    "predana_energija":  "Predaja u el. mrežu (kWh)",
    "preuzeta_energija": "Preuzeta energija iz mreže (kWh)",
    "razlika_energija":  "Razlika između potrošnje i proizvodnje (kWh)",
}

PDF_FIELD_HINTS = {
    "ime": "ime_prezime", "naziv": "ime_prezime",
    "oib":  "oib",
    "ulica": "inv_ulica", "ulica_naziv": "inv_ulica",
    "kucni": "inv_broj",  "broj": "inv_broj",
    "postal": "inv_post", "postansk": "inv_post",
    "grad": "inv_grad",   "mjesto": "inv_grad",
    "brm": "brm", "omm": "brm",
    "zakup": "zak_snaga",
    "dc_snaga": "dc_snaga", "ac_snaga": "ac_snaga",
    "br_pan": "br_panela",  "snaga_pan": "snaga_panela",
    "pr_pan": "pr_pan",     "md_pan": "md_pan",
    "br_inv": "br_inv",     "snaga_inv": "snaga_inv",
    "pr_inv": "pr_inv",     "md_inv": "md_inv",
    "datum": "mj_dat",
    "tel": "telefon",  "mob": "telefon",
    "email": "email",  "mail": "email",
    "cestica": "kat_cestica", "opcina": "kat_opcina",
}

# ---------------------------------------------------------------------------
# Default rasporedi polja za poznate obrasce
# Pozicije su u PDF točkama (pt).
# Y = vrh teksta (generate_overlay dodaje font_size za baseline).
# ---------------------------------------------------------------------------
# Koordinate dobivene analizom stvarnih PDF predložaka.

def _f(x, y, page=0, fs=10):
    return {"x": x, "y": y, "font_size": fs, "page": page}

DEFAULT_LAYOUTS = {
    # Ključ: string koji mora biti substring u basename(pdf_path).lower()
    "pm1.2.1": {
        "fields": {
            # ── INVESTITOR (stranica 0) ───────────────────────────────
            "ime_prezime":  _f(180,  121.5),
            "oib":          _f(430,  121.5),
            "inv_grad":     _f(133,  139.1),
            "inv_post":     _f(346,  139.1),
            "inv_ulica":    _f(126,  156.6),
            "inv_broj":     _f(299,  156.6),
            "email":        _f(290,  174.0),
            "telefon":      _f(467,  174.0),
            # ── OPUNOMOĆENIK / PROJEKTANT ────────────────────────────
            "projektant":   _f(180,  214.5),
            # ── PODACI O GRAĐEVINI (dno str. 0) ─────────────────────
            "pm_grad":      _f(161,  715.5),
            "pm_post":      _f(374,  715.5),
            "pm_ulica":     _f(153,  732.9),
            "pm_broj":      _f(327,  732.9),
            "kat_cestica":  _f(119,  750.4),
            "kat_opcina":   _f(341,  750.4),
        }
    },
    "pm1.6.1": {
        "fields": {
            # ── VLASNIK GRAĐEVINE (stranica 0) ───────────────────────
            "ime_prezime":  _f(180,  127.5),
            "oib":          _f(430,  127.5),
            "inv_grad":     _f(133,  145.1),
            "inv_post":     _f(346,  145.1),
            "inv_ulica":    _f(126,  162.6),
            "inv_broj":     _f(299,  162.6),
            "email":        _f(290,  180.1),
            "telefon":      _f(467,  180.1),
            # ── OPUNOMOĆENIK ─────────────────────────────────────────
            "projektant":   _f(180,  232.5),
            # ── PODACI O GRAĐEVINI ────────────────────────────────────
            "pm_grad":      _f(161,  432.0),
            "pm_post":      _f(374,  432.0),
            "pm_ulica":     _f(153,  449.5),
            "pm_broj":      _f(327,  449.5),
            "kat_cestica":  _f(119,  466.9),
            "kat_opcina":   _f(341,  466.9),
            # ── PODACI O POSTOJEĆEM OMM ──────────────────────────────
            "brm":          _f(133,  749.9),
            "zak_snaga":    _f(460,  749.9),
            # ── PRIKLJUČNA SNAGA ─────────────────────────────────────
            "ac_snaga":     _f(202,  501.9),
        }
    },
}


def get_default_layout(pdf_path: str) -> dict | None:
    """Vraća default layout za poznate obrasce, None ako nema podudaranja."""
    base = os.path.basename(pdf_path).lower().replace("-", "").replace("_", "").replace(".", "")
    for key, layout in DEFAULT_LAYOUTS.items():
        if key.replace(".", "") in base:
            return json.loads(json.dumps(layout))   # deep copy
    return None


# ---------------------------------------------------------------------------
# Pomoćne funkcije
# ---------------------------------------------------------------------------

def layout_path(pdf_path: str) -> str:
    return os.path.splitext(pdf_path)[0] + LAYOUT_SUFFIX


def load_layout(pdf_path: str) -> dict:
    lp = layout_path(pdf_path)
    if os.path.exists(lp):
        try:
            with open(lp, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    # Ako nema layout datoteke, pokušaj default
    default = get_default_layout(pdf_path)
    return default if default else {"fields": {}}


def save_layout(pdf_path: str, layout: dict):
    with open(layout_path(pdf_path), "w", encoding="utf-8") as f:
        json.dump(layout, f, ensure_ascii=False, indent=2)


def open_file(path: str):
    if os.name == "nt":
        os.startfile(path)
    else:
        subprocess.run(["xdg-open", path])


def post_process(data: dict) -> dict:
    """Dodaje split adresna polja derivirana iz sirovih Excel vrijednosti."""
    result = dict(data)

    # adresa_ulica → inv_ulica + inv_broj
    if "adresa_ulica" in data:
        addr = data["adresa_ulica"].strip()
        m = re.match(r"^(.+?)\s+(\d\S*)$", addr)
        if m:
            result["inv_ulica"] = m.group(1)
            result["inv_broj"]  = m.group(2)
        else:
            result["inv_ulica"] = addr
            result["inv_broj"]  = ""

    # adresa_grad → inv_post + inv_grad
    if "adresa_grad" in data:
        grad = data["adresa_grad"].strip()
        m = re.match(r"^(\d{5})\s+(.+)$", grad)
        if m:
            result["inv_post"] = m.group(1)
            result["inv_grad"] = m.group(2)
        else:
            result["inv_post"] = ""
            result["inv_grad"] = grad

    # adresa_pm → pm_ulica + pm_broj + pm_post + pm_grad
    if "adresa_pm" in data:
        pm = data["adresa_pm"].strip().lstrip("\xa0")
        parts = [p.strip() for p in pm.split(",")]
        if len(parts) >= 2:
            m = re.match(r"^(.+?)\s+(\d\S*)$", parts[0])
            if m:
                result["pm_ulica"] = m.group(1)
                result["pm_broj"]  = m.group(2)
            else:
                result["pm_ulica"] = parts[0]
                result["pm_broj"]  = ""
            m2 = re.match(r"^(\d{5})\s+(.+)$", parts[1])
            if m2:
                result["pm_post"] = m2.group(1)
                result["pm_grad"] = m2.group(2)
            else:
                result["pm_post"] = ""
                result["pm_grad"] = parts[1]
        else:
            result.setdefault("pm_ulica", pm)
            result.setdefault("pm_broj",  "")
            result.setdefault("pm_post",  "")
            result.setdefault("pm_grad",  "")

    # lokacija → kat_cestica + kat_opcina
    if "lokacija" in data:
        lok = data["lokacija"]
        m = re.search(
            r"K\.Č\.\s*([^,]+),\s*K\.O\.\s*(.+)", lok, re.IGNORECASE)
        if m:
            result["kat_cestica"] = m.group(1).strip()
            result["kat_opcina"]  = m.group(2).strip()

    return result


def extract_from_excel(path: str) -> dict:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = next(
            (n for n in wb.sheetnames if "podaci" in n.lower()), None)
        if sheet is None:
            return {}
        ws = wb[sheet]
        found = {}
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if not isinstance(val, str):
                    continue
                t = val.strip().lower()
                for kw, tag in EXCEL_KEYWORDS.items():
                    if kw in t and tag not in found:
                        nxt = ws.cell(
                            row=cell.row, column=cell.column + 1).value
                        if nxt is not None and str(nxt).strip():
                            found[tag] = str(nxt).strip()
        return post_process(found)
    finally:
        wb.close()


def get_form_fields(pdf_path: str) -> dict:
    if not HAS_PYPDFFORM:
        return {}
    try:
        w = PdfWrapper(pdf_path)
        return dict(w.data) if w.data else {}
    except Exception:
        return {}


def auto_map_form_fields(form_fields: dict, excel_data: dict) -> dict:
    out = {}
    for field in form_fields:
        fl = field.lower().replace("-", "_").replace(" ", "_")
        for hint, tag in PDF_FIELD_HINTS.items():
            if hint in fl and tag in excel_data:
                out[field] = excel_data[tag]
                break
    return out


# ---------------------------------------------------------------------------
# PDF generiranje
# ---------------------------------------------------------------------------

def generate_overlay(pdf_path: str, layout: dict, excel_data: dict,
                     output_path: str):
    """Overlay tekst na PDF koristeći Unicode TrueType font.
    cfg['y'] je vrh teksta; dodajemo font_size*0.85 za fitz baseline."""
    global _font_file
    doc = fitz.open(pdf_path)

    # Grupiraj polja po stranicama (registracija fonta per-page)
    by_page: dict[int, list] = {}
    for tag, cfg in layout.get("fields", {}).items():
        pg = int(cfg.get("page", 0))
        by_page.setdefault(pg, []).append((tag, cfg))

    for page_i, fields in by_page.items():
        if page_i >= len(doc):
            continue
        page = doc[page_i]
        # Registriraj font jednom po stranici
        font_kwargs: dict = {}
        if _font_file and os.path.exists(_font_file):
            font_kwargs = {"fontname": "F0", "fontfile": _font_file}

        for tag, cfg in fields:
            value = excel_data.get(tag) or cfg.get("manual_value", "")
            if not value:
                continue
            fs         = int(cfg.get("font_size", 10))
            baseline_y = float(cfg["y"]) + fs * 0.85
            page.insert_text(
                fitz.Point(float(cfg["x"]), baseline_y),
                str(value),
                fontsize=fs,
                color=(0, 0, 0),
                **font_kwargs,
            )

    doc.save(output_path)
    doc.close()


def generate_form(pdf_path: str, field_values: dict, output_path: str):
    w     = PdfWrapper(pdf_path)
    valid = {k: v for k, v in field_values.items() if k in (w.data or {})}
    out   = w.fill(valid)
    with open(output_path, "wb") as f:
        f.write(out.read())


# ---------------------------------------------------------------------------
# Glavna aplikacija — jedan prozor
# ---------------------------------------------------------------------------

class HepaApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("HEPA Form Filler")
        self.root.geometry("1280x820")
        self.root.resizable(True, True)

        self.excel_path = tk.StringVar()
        self.pdf_path   = tk.StringVar()
        self.output_dir = tk.StringVar()

        self.excel_data:  dict = {}
        self.layout:      dict = {"fields": {}}
        self.form_fields: dict = {}

        # Editor state
        self._doc:           fitz.Document | None = None
        self._page_count   = 0
        self._current_page = 0
        self._scale        = 1.0          # canvas px / PDF pt
        self._img          = None         # PhotoImage (anti-GC)
        self._selected     = None
        self._pick_mode    = False
        self._drag         = {"active": False, "last_x": 0.0,
                              "last_y": 0.0, "tag": None}
        self._fields: dict = {}
        self._manual_updating = False
        self._custom_counter  = 0

        self._build_ui()

    # ------------------------------------------------------------------
    # Pomoćne UI metode
    # ------------------------------------------------------------------

    def _get_label(self, tag: str, cfg: dict = None) -> str:
        if cfg is None:
            cfg = self._fields.get(tag, {})
        return cfg.get("label") or TAG_LABELS.get(tag, tag)

    def _sync_custom_counter(self):
        """Ažurira _custom_counter da izbjegne kolizije pri dodavanju novih polja."""
        for tag in self._fields:
            if tag.startswith("custom__"):
                try:
                    n = int(tag.split("__", 1)[1])
                    self._custom_counter = max(self._custom_counter, n + 1)
                except (ValueError, IndexError):
                    pass

    # ------------------------------------------------------------------
    # Gradnja UI
    # ------------------------------------------------------------------

    def _build_ui(self):
        # ── Red 1: odabir datoteka ───────────────────────────────────────
        row1 = ttk.Frame(self.root, relief="raised")
        row1.pack(fill="x")

        ttk.Label(row1, text="Excel:").pack(side="left", padx=(8, 2), pady=4)
        ttk.Entry(row1, textvariable=self.excel_path,
                  width=36).pack(side="left", pady=4)
        ttk.Button(row1, text="…", width=2,
                   command=self._pick_excel).pack(side="left", padx=(0, 8))

        ttk.Label(row1, text="PDF:").pack(side="left", padx=(0, 2))
        ttk.Entry(row1, textvariable=self.pdf_path,
                  width=36).pack(side="left", pady=4)
        ttk.Button(row1, text="…", width=2,
                   command=self._pick_pdf).pack(side="left", padx=(0, 8))

        ttk.Label(row1, text="Spremi u:").pack(side="left", padx=(0, 2))
        ttk.Entry(row1, textvariable=self.output_dir,
                  width=24).pack(side="left", pady=4)
        ttk.Button(row1, text="…", width=2,
                   command=self._pick_outdir).pack(side="left")

        # ── Red 2: akcije + zoom + stranice ─────────────────────────────
        row2 = ttk.Frame(self.root, relief="raised")
        row2.pack(fill="x")

        ttk.Button(row2, text="📂  Učitaj",
                   command=self._load).pack(side="left", padx=(8, 4), pady=3)
        ttk.Button(row2, text="📄  Generiraj PDF",
                   command=self._generate).pack(side="left", padx=4)

        ttk.Separator(row2, orient="vertical").pack(
            side="left", fill="y", padx=10, pady=3)

        ttk.Label(row2, text="Font:").pack(side="left", padx=(0, 2))
        self._font_lbl = ttk.Label(
            row2,
            text=os.path.basename(_font_file) if _font_file else "⚠ nije pronađen",
            foreground="#1a5276" if _font_file else "#c0392b",
            width=22)
        self._font_lbl.pack(side="left")
        ttk.Button(row2, text="…", width=2,
                   command=self._pick_font).pack(side="left", padx=(0, 8))

        ttk.Separator(row2, orient="vertical").pack(
            side="left", fill="y", padx=10, pady=3)

        ttk.Label(row2, text="Raspored:").pack(side="left", padx=(0, 4))
        ttk.Button(row2, text="💾  Spremi",
                   command=self._save_layout).pack(side="left", padx=2)
        ttk.Button(row2, text="📂  Otvori JSON",
                   command=self._load_layout_from_file).pack(side="left", padx=2)
        ttk.Button(row2, text="↺  Default",
                   command=self._load_default_layout).pack(side="left", padx=2)

        ttk.Separator(row2, orient="vertical").pack(
            side="left", fill="y", padx=10, pady=3)

        ttk.Label(row2, text="Zoom:").pack(side="left")
        ttk.Button(row2, text="−", width=2,
                   command=self._zoom_out).pack(side="left", padx=2)
        ttk.Button(row2, text="fit", width=3,
                   command=self._zoom_fit).pack(side="left", padx=2)
        ttk.Button(row2, text="+", width=2,
                   command=self._zoom_in).pack(side="left", padx=2)

        ttk.Separator(row2, orient="vertical").pack(
            side="left", fill="y", padx=10, pady=3)

        ttk.Button(row2, text="▶", width=2,
                   command=self._next_page).pack(side="right", padx=(2, 8))
        self._page_lbl = ttk.Label(row2, text="—", width=7, anchor="center")
        self._page_lbl.pack(side="right")
        ttk.Button(row2, text="◀", width=2,
                   command=self._prev_page).pack(side="right", padx=2)
        ttk.Label(row2, text="Str:").pack(side="right", padx=(0, 2))

        # ── Glavni panel: lijevo + desno ─────────────────────────────────
        paned = ttk.PanedWindow(self.root, orient="horizontal")
        paned.pack(fill="both", expand=True)

        # -- Lijevi panel ------------------------------------------------
        left = ttk.Frame(paned, width=300)
        left.pack_propagate(False)
        paned.add(left, weight=0)

        ef = ttk.LabelFrame(left, text="Učitani Excel podaci")
        ef.pack(fill="both", expand=True, padx=4, pady=(4, 2))
        self.excel_tree = ttk.Treeview(
            ef, columns=("polje", "vrijednost"), show="headings", height=10)
        self.excel_tree.heading("polje",      text="Polje")
        self.excel_tree.heading("vrijednost", text="Vrijednost")
        self.excel_tree.column("polje",      width=130, stretch=True)
        self.excel_tree.column("vrijednost", width=140, stretch=True)
        tsb = ttk.Scrollbar(ef, orient="vertical",
                            command=self.excel_tree.yview)
        self.excel_tree.configure(yscrollcommand=tsb.set)
        self.excel_tree.pack(side="left", fill="both", expand=True)
        tsb.pack(side="right", fill="y")

        pf = ttk.LabelFrame(left, text="Polja na obrascu")
        pf.pack(fill="x", padx=4, pady=2)
        lf = ttk.Frame(pf)
        lf.pack(fill="x", padx=2, pady=2)
        self._listbox = tk.Listbox(lf, selectmode="single", font=("", 9),
                                   height=7, activestyle="dotbox")
        lsb = ttk.Scrollbar(lf, orient="vertical",
                            command=self._listbox.yview)
        self._listbox.configure(yscrollcommand=lsb.set)
        self._listbox.pack(side="left", fill="both", expand=True)
        lsb.pack(side="right", fill="y")
        self._listbox.bind("<<ListboxSelect>>", self._on_list_select)

        af = ttk.Frame(pf)
        af.pack(fill="x", padx=4, pady=2)
        self._new_tag_var = tk.StringVar()
        self._add_tags    = sorted(TAG_LABELS, key=lambda t: TAG_LABELS[t])
        self._add_combo   = ttk.Combobox(
            af, textvariable=self._new_tag_var,
            values=[TAG_LABELS[t] for t in self._add_tags],
            state="readonly", width=26)
        self._add_combo.pack(side="left", padx=(0, 4))
        ttk.Button(af, text="＋", width=3,
                   command=self._add_field).pack(side="left")

        cf = ttk.Frame(pf)
        cf.pack(fill="x", padx=4, pady=(0, 4))
        ttk.Label(cf, text="Vlastito:", width=7).pack(side="left")
        self._custom_name_var = tk.StringVar()
        ttk.Entry(cf, textvariable=self._custom_name_var,
                  width=18).pack(side="left", padx=(0, 4))
        ttk.Button(cf, text="＋", width=3,
                   command=self._add_custom_field).pack(side="left")

        sf = ttk.LabelFrame(left, text="Odabrano polje")
        sf.pack(fill="x", padx=4, pady=4)

        self._sel_name  = ttk.Label(sf, text="—", font=("", 9, "bold"))
        self._sel_name.pack(anchor="w", padx=6, pady=(4, 0))
        self._sel_value = ttk.Label(sf, text="", foreground="#1a5276",
                                    wraplength=270, justify="left")
        self._sel_value.pack(anchor="w", padx=6, pady=(0, 4))
        ttk.Separator(sf).pack(fill="x", padx=4)

        ttk.Label(sf, text="Ručna vrijednost:").pack(
            anchor="w", padx=6, pady=(4, 0))
        self._manual_var = tk.StringVar()
        ttk.Entry(sf, textvariable=self._manual_var,
                  width=34).pack(padx=6, pady=2, fill="x")
        self._manual_var.trace_add("write", self._on_manual_change)
        ttk.Separator(sf).pack(fill="x", padx=4, pady=4)

        ttk.Label(sf, text="Veličina fonta (pt):").pack(
            anchor="w", padx=6)
        ff = ttk.Frame(sf)
        ff.pack(padx=6, pady=2, fill="x")
        ttk.Button(ff, text="−", width=3,
                   command=lambda: self._change_font(-1)).pack(side="left")
        self._font_var = tk.IntVar(value=10)
        ttk.Spinbox(ff, textvariable=self._font_var, from_=5, to=48,
                    width=5, command=self._apply_font).pack(
            side="left", padx=4)
        ttk.Button(ff, text="+", width=3,
                   command=lambda: self._change_font(+1)).pack(side="left")
        ttk.Label(ff, text=" scroll = promjena",
                  foreground="#777").pack(side="left")

        ttk.Separator(sf).pack(fill="x", padx=4, pady=4)
        self._pick_btn = ttk.Button(
            sf, text="📍  Pozicija — klikni na PDF",
            command=self._enable_pick)
        self._pick_btn.pack(fill="x", padx=6, pady=2)
        ttk.Button(sf, text="🗑  Ukloni polje",
                   command=self._remove_field).pack(
            fill="x", padx=6, pady=(2, 6))

        # -- Desni panel: PDF canvas ------------------------------------
        right = ttk.Frame(paned)
        paned.add(right, weight=1)

        self.canvas = tk.Canvas(right, bg="#606060", cursor="arrow",
                                highlightthickness=0)
        h_sb = ttk.Scrollbar(right, orient="horizontal",
                             command=self.canvas.xview)
        v_sb = ttk.Scrollbar(right, orient="vertical",
                             command=self.canvas.yview)
        self.canvas.configure(xscrollcommand=h_sb.set,
                              yscrollcommand=v_sb.set)
        h_sb.pack(side="bottom", fill="x")
        v_sb.pack(side="right",  fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.canvas.bind("<Button-1>",       self._cv_click)
        self.canvas.bind("<B1-Motion>",       self._cv_drag)
        self.canvas.bind("<ButtonRelease-1>", self._cv_release)
        self.canvas.bind("<MouseWheel>",      self._cv_scroll)
        self.canvas.bind("<Button-4>",        self._cv_scroll)
        self.canvas.bind("<Button-5>",        self._cv_scroll)

        # ── Status bar ──────────────────────────────────────────────────
        self._status = tk.StringVar(
            value="Odaberite Excel i PDF, zatim kliknite 'Učitaj'.")
        ttk.Label(self.root, textvariable=self._status,
                  relief="sunken", anchor="w").pack(
            fill="x", side="bottom", ipady=2)

    # ------------------------------------------------------------------
    # Browse callbacks
    # ------------------------------------------------------------------

    def _pick_excel(self):
        p = filedialog.askopenfilename(
            title="Odaberite Excel projektnu datoteku",
            filetypes=[("Excel", "*.xlsm *.xlsx"), ("Sve", "*.*")])
        if p:
            self.excel_path.set(p)
            self.output_dir.set(os.path.dirname(p))

    def _pick_pdf(self):
        p = filedialog.askopenfilename(
            title="Odaberite HEPA PDF predložak",
            filetypes=[("PDF", "*.pdf"), ("Sve", "*.*")])
        if p:
            self.pdf_path.set(p)

    def _pick_outdir(self):
        p = filedialog.askdirectory(title="Izlazna mapa")
        if p:
            self.output_dir.set(p)

    def _pick_font(self):
        global _font_file
        p = filedialog.askopenfilename(
            title="Odaberite TrueType font (.ttf)",
            filetypes=[("TrueType font", "*.ttf *.otf"), ("Sve", "*.*")])
        if p:
            _font_file = p
            self._font_lbl.config(
                text=os.path.basename(p), foreground="#1a5276")
            self._status.set(f"Font postavljen: {p}")

    # ------------------------------------------------------------------
    # Učitavanje
    # ------------------------------------------------------------------

    def _load(self):
        if not self.excel_path.get() or not self.pdf_path.get():
            messagebox.showwarning("Upozorenje", "Odaberite Excel i PDF!")
            return

        self._status.set("Učitavam…")
        self.root.update_idletasks()

        try:
            self.excel_data = extract_from_excel(self.excel_path.get())
        except Exception as e:
            messagebox.showerror("Greška", f"Excel greška:\n{e}")
            self._status.set("Greška.")
            return

        # Prikaži Excel podatke (bez internih kombinirani tagova)
        display_skip = {"adresa_ulica", "adresa_grad", "adresa_pm",
                        "lokacija"}
        self.excel_tree.delete(*self.excel_tree.get_children())
        for tag, val in self.excel_data.items():
            if tag in display_skip:
                continue
            self.excel_tree.insert("", "end",
                                   values=(TAG_LABELS.get(tag, tag), val))

        self.form_fields = get_form_fields(self.pdf_path.get())
        self.layout      = load_layout(self.pdf_path.get())
        self._fields     = {t: dict(c)
                            for t, c in self.layout.get("fields", {}).items()}
        self._sync_custom_counter()

        if self._doc:
            self._doc.close()
        self._doc          = fitz.open(self.pdf_path.get())
        self._page_count   = len(self._doc)
        self._current_page = 0

        # Računaj scale da odgovara širini canvasa
        self.root.update_idletasks()
        self._scale = self._fit_scale()

        self._render_page()
        self._draw_all_fields()
        self._refresh_listbox()

        used_default = not os.path.exists(layout_path(self.pdf_path.get()))
        pdf_type = "fillable" if self.form_fields else "flat/overlay"
        n_layout = len(self._fields)
        hint = (" [default raspored učitan]" if used_default and n_layout
                else "" if n_layout
                else "  ← dodajte polja s lijeva")
        self._status.set(
            f"Excel: {len(self.excel_data)} polja  |  "
            f"PDF: {pdf_type}  |  Raspored: {n_layout} polja{hint}")

    # ------------------------------------------------------------------
    # Renderiranje i zoom
    # ------------------------------------------------------------------

    def _fit_scale(self) -> float:
        """Izračuna scale koji uklapa PDF u trenutnu širinu canvasa."""
        if not self._doc:
            return 1.0
        self.canvas.update_idletasks()
        cw = max(200, self.canvas.winfo_width())
        pw = self._doc[self._current_page].rect.width
        return min(MAX_SCALE, max(MIN_SCALE, (cw - 20) / pw))

    def _render_page(self):
        if not self._doc:
            return
        page = self._doc[self._current_page]
        mat  = fitz.Matrix(self._scale, self._scale)
        pix  = page.get_pixmap(matrix=mat, alpha=False)
        img  = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        self._img = ImageTk.PhotoImage(img)
        self.canvas.delete("all")
        self.canvas.create_image(0, 0, anchor="nw", image=self._img,
                                 tags="bg")
        self.canvas.configure(
            scrollregion=(0, 0, pix.width, pix.height))
        self._page_lbl.config(
            text=f"{self._current_page + 1} / {self._page_count}")

    def _zoom_in(self):
        self._scale = min(MAX_SCALE, self._scale * 1.25)
        self._refresh_view()

    def _zoom_out(self):
        self._scale = max(MIN_SCALE, self._scale * 0.8)
        self._refresh_view()

    def _zoom_fit(self):
        self._scale = self._fit_scale()
        self._refresh_view()

    def _refresh_view(self):
        self._render_page()
        self._draw_all_fields()

    def _prev_page(self):
        if self._doc and self._current_page > 0:
            self._current_page -= 1
            self._render_page()
            self._draw_all_fields()
            self._refresh_listbox()

    def _next_page(self):
        if self._doc and self._current_page < self._page_count - 1:
            self._current_page += 1
            self._render_page()
            self._draw_all_fields()
            self._refresh_listbox()

    # ------------------------------------------------------------------
    # Crtanje polja
    # ------------------------------------------------------------------

    def _draw_all_fields(self):
        for tag, cfg in self._fields.items():
            if int(cfg.get("page", 0)) == self._current_page:
                self._draw_field(tag)

    def _draw_field(self, tag: str):
        cfg     = self._fields[tag]
        cx      = cfg["x"] * self._scale
        cy      = cfg["y"] * self._scale
        value   = (self.excel_data.get(tag)
                   or cfg.get("manual_value", "")
                   or f"[{self._get_label(tag, cfg)}]")
        canvas_fs = max(6, int(cfg["font_size"] * self._scale))
        is_sel    = (tag == self._selected)
        fill_col  = "#c0392b" if is_sel else "#111111"
        ctag      = f"field_{tag}"

        self.canvas.delete(ctag)
        text_id = self.canvas.create_text(
            cx, cy, text=str(value), font=("Arial", canvas_fs),
            fill=fill_col, anchor="nw", tags=(ctag, "field_item"))

        bbox = self.canvas.bbox(text_id)
        if bbox:
            p = 4
            self.canvas.create_rectangle(
                bbox[0]-p, bbox[1]-p, bbox[2]+p, bbox[3]+p,
                fill="", outline="#c0392b" if is_sel else "",
                width=1 if is_sel else 0,
                tags=(ctag, "field_item", "hitbox"))

    # ------------------------------------------------------------------
    # Canvas događaji
    # ------------------------------------------------------------------

    def _cv_click(self, event):
        cx = self.canvas.canvasx(event.x)
        cy = self.canvas.canvasy(event.y)

        if self._pick_mode and self._selected:
            self._fields[self._selected]["x"] = cx / self._scale
            self._fields[self._selected]["y"] = cy / self._scale
            self._draw_field(self._selected)
            self._pick_mode = False
            self.canvas.config(cursor="arrow")
            self._pick_btn.config(
                text="📍  Pozicija — klikni na PDF")
            self._status.set(
                f"Pozicija postavljena: "
                f"{self._get_label(self._selected)}")
            return

        found = self._field_at(cx, cy)
        if found:
            self._select(found)
            self._drag.update(active=True, last_x=cx, last_y=cy, tag=found)
        else:
            self._deselect()

    def _cv_drag(self, event):
        if not self._drag["active"] or not self._drag["tag"]:
            return
        cx  = self.canvas.canvasx(event.x)
        cy  = self.canvas.canvasy(event.y)
        dx  = cx - self._drag["last_x"]
        dy  = cy - self._drag["last_y"]
        tag = self._drag["tag"]

        # Pomakni canvas item
        self.canvas.move(f"field_{tag}", dx, dy)

        # Ažuriraj poziciju u PDF točkama
        self._fields[tag]["x"] += dx / self._scale
        self._fields[tag]["y"] += dy / self._scale

        self._drag["last_x"] = cx
        self._drag["last_y"] = cy

    def _cv_release(self, event):
        if self._drag["tag"]:
            # Redraw da hitbox prati novu poziciju
            self._draw_field(self._drag["tag"])
        self._drag.update(active=False, tag=None)

    def _cv_scroll(self, event):
        if not self._selected:
            return
        up = (getattr(event, "num", 0) == 4
              or getattr(event, "delta", 0) > 0)
        self._change_font(+1 if up else -1)

    def _field_at(self, cx: float, cy: float) -> "str | None":
        for iid in reversed(
                self.canvas.find_overlapping(cx-6, cy-6, cx+6, cy+6)):
            for t in self.canvas.gettags(iid):
                if t.startswith("field_"):
                    return t[6:]
        return None

    # ------------------------------------------------------------------
    # Selekcija
    # ------------------------------------------------------------------

    def _select(self, tag: str):
        prev = self._selected
        self._selected = tag
        if prev and prev != tag:
            self._draw_field(prev)
        self._draw_field(tag)
        self._update_controls()
        tags_pg = [t for t, c in self._fields.items()
                   if int(c.get("page", 0)) == self._current_page]
        if tag in tags_pg:
            idx = tags_pg.index(tag)
            self._listbox.selection_clear(0, "end")
            self._listbox.selection_set(idx)
            self._listbox.see(idx)

    def _deselect(self):
        prev, self._selected = self._selected, None
        if prev:
            self._draw_field(prev)
        self._update_controls()
        self._listbox.selection_clear(0, "end")

    def _update_controls(self):
        tag = self._selected
        if tag is None:
            self._sel_name.config(text="—")
            self._sel_value.config(text="")
            self._manual_updating = True
            self._manual_var.set("")
            self._manual_updating = False
            return
        self._sel_name.config(text=self._get_label(tag))
        if tag.startswith("custom__"):
            self._sel_value.config(text="(vlastito polje — upiši ručnu vrijednost)")
        else:
            val = self.excel_data.get(tag, "")
            self._sel_value.config(
                text=f"Excel: {val}" if val else "(nema u Excelu)")
        self._font_var.set(int(self._fields[tag].get("font_size", 10)))
        self._manual_updating = True
        self._manual_var.set(self._fields[tag].get("manual_value", ""))
        self._manual_updating = False

    # ------------------------------------------------------------------
    # Listbox
    # ------------------------------------------------------------------

    def _refresh_listbox(self):
        self._listbox.delete(0, "end")
        for tag, cfg in self._fields.items():
            if int(cfg.get("page", 0)) != self._current_page:
                continue
            val  = (self.excel_data.get(tag)
                    or cfg.get("manual_value", ""))
            tick = "✓" if val else "○"
            self._listbox.insert(
                "end", f"{tick}  {self._get_label(tag, cfg)}")

    def _on_list_select(self, _event=None):
        sel = self._listbox.curselection()
        if not sel:
            return
        tags_pg = [t for t, c in self._fields.items()
                   if int(c.get("page", 0)) == self._current_page]
        idx = sel[0]
        if idx < len(tags_pg):
            self._select(tags_pg[idx])

    # ------------------------------------------------------------------
    # Dodaj / ukloni polje
    # ------------------------------------------------------------------

    def _add_field(self):
        label = self._new_tag_var.get()
        if not label:
            return
        tag = next((t for t in self._add_tags
                    if TAG_LABELS[t] == label), None)
        if tag is None:
            return
        if (tag in self._fields and
                int(self._fields[tag].get("page", 0)) == self._current_page):
            messagebox.showinfo("Info",
                                "To polje već postoji na ovoj stranici.")
            return
        if self._doc:
            r = self._doc[self._current_page].rect
            self._fields[tag] = {
                "x": r.width / 2, "y": r.height / 2,
                "font_size": 10, "page": self._current_page}
        else:
            self._fields[tag] = {"x": 100, "y": 100,
                                  "font_size": 10, "page": 0}
        self._draw_field(tag)
        self._refresh_listbox()
        self._select(tag)
        self._status.set(
            f"Dodano: {self._get_label(tag)}. "
            "Drag za pomicanje ili '📍 Pozicija'.")

    def _add_custom_field(self):
        name = self._custom_name_var.get().strip()
        if not name:
            messagebox.showwarning("Upozorenje", "Unesite naziv polja!")
            return
        tag = f"custom__{self._custom_counter}"
        self._custom_counter += 1
        if self._doc:
            r = self._doc[self._current_page].rect
            self._fields[tag] = {
                "label": name,
                "x": r.width / 2, "y": r.height / 2,
                "font_size": 10, "page": self._current_page,
            }
        else:
            self._fields[tag] = {
                "label": name, "x": 100, "y": 100,
                "font_size": 10, "page": 0,
            }
        self._custom_name_var.set("")
        self._draw_field(tag)
        self._refresh_listbox()
        self._select(tag)
        self._status.set(
            f"Dodano vlastito polje: {name}. Upiši vrijednost i postavi poziciju.")

    def _remove_field(self):
        if not self._selected:
            return
        tag = self._selected
        label = self._get_label(tag)
        self.canvas.delete(f"field_{tag}")
        del self._fields[tag]
        self._selected = None
        self._refresh_listbox()
        self._update_controls()
        self._status.set(f"Uklonjeno: {label}.")

    # ------------------------------------------------------------------
    # Font
    # ------------------------------------------------------------------

    def _change_font(self, delta: int):
        if not self._selected:
            return
        fs = max(5, int(self._fields[self._selected].get(
            "font_size", 10)) + delta)
        self._fields[self._selected]["font_size"] = fs
        self._font_var.set(fs)
        self._draw_field(self._selected)

    def _apply_font(self):
        if not self._selected:
            return
        self._fields[self._selected]["font_size"] = self._font_var.get()
        self._draw_field(self._selected)

    # ------------------------------------------------------------------
    # Pick-mode
    # ------------------------------------------------------------------

    def _enable_pick(self):
        if not self._selected:
            messagebox.showwarning("Upozorenje", "Najprije odaberite polje!")
            return
        self._pick_mode = True
        self.canvas.config(cursor="crosshair")
        self._pick_btn.config(text="⌛  Kliknite na PDF za poziciju…")
        self._status.set("Pick mode — kliknite na PDF gdje želite tekst.")

    # ------------------------------------------------------------------
    # Ručna vrijednost
    # ------------------------------------------------------------------

    def _on_manual_change(self, *_):
        if not self._selected or self._manual_updating:
            return
        self._fields[self._selected]["manual_value"] = self._manual_var.get()
        self._draw_field(self._selected)
        self._refresh_listbox()

    # ------------------------------------------------------------------
    # Spremi raspored
    # ------------------------------------------------------------------

    def _save_layout(self):
        if not self.pdf_path.get():
            messagebox.showwarning("Upozorenje", "Nema učitanog PDF-a!")
            return
        self.layout["fields"] = {
            tag: {k: v for k, v in cfg.items()}
            for tag, cfg in self._fields.items()
        }
        save_layout(self.pdf_path.get(), self.layout)
        self._status.set(
            f"Raspored spremljen: {layout_path(self.pdf_path.get())}")

    def _load_layout_from_file(self):
        """Otvori bilo koji _layout.json i primijeni ga."""
        if not self._doc:
            messagebox.showwarning("Upozorenje", "Najprije učitajte PDF!")
            return
        p = filedialog.askopenfilename(
            title="Odaberite raspored (JSON)",
            initialdir=(os.path.dirname(self.pdf_path.get())
                        if self.pdf_path.get() else "."),
            filetypes=[("JSON raspored", "*.json"), ("Sve", "*.*")])
        if not p:
            return
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("Greška", f"Ne mogu čitati JSON:\n{e}")
            return
        self._apply_layout(data, label=os.path.basename(p))

    def _load_default_layout(self):
        """Primijeni ugrađeni default raspored za trenutni PDF."""
        if not self.pdf_path.get():
            messagebox.showwarning("Upozorenje", "Najprije učitajte PDF!")
            return
        default = get_default_layout(self.pdf_path.get())
        if default is None:
            messagebox.showinfo(
                "Info",
                "Nema ugrađenog default rasporeda za ovaj PDF.\n"
                "Podržani obrasci: PM1.2.1, PM1.6.1")
            return
        self._apply_layout(default, label="default")

    def _apply_layout(self, layout: dict, label: str = ""):
        """Zamijeni trenutna polja onima iz layouta i osvježi prikaz."""
        # Obriši stara polja s canvasa
        for tag in list(self._fields):
            self.canvas.delete(f"field_{tag}")

        self.layout  = layout
        self._fields = {t: dict(c)
                        for t, c in layout.get("fields", {}).items()}
        self._sync_custom_counter()
        self._selected = None
        self._update_controls()
        self._render_page()
        self._draw_all_fields()
        self._refresh_listbox()
        n = len(self._fields)
        self._status.set(
            f"Raspored učitan ({label}): {n} polja definiranih.")

    # ------------------------------------------------------------------
    # Generiranje PDF-a
    # ------------------------------------------------------------------

    def _generate(self):
        if not self.excel_data and not self._fields:
            messagebox.showwarning("Upozorenje",
                                   "Učitajte podatke i postavite polja!")
            return
        if not self.output_dir.get():
            messagebox.showwarning("Upozorenje", "Odaberite izlaznu mapu!")
            return

        pdf_base   = os.path.splitext(
            os.path.basename(self.pdf_path.get()))[0]
        excel_base = os.path.splitext(
            os.path.basename(self.excel_path.get()))[0]
        out_path   = os.path.join(
            self.output_dir.get(), f"{pdf_base}__{excel_base}.pdf")

        self.layout["fields"] = {
            tag: {k: v for k, v in cfg.items()}
            for tag, cfg in self._fields.items()
        }

        try:
            if self.form_fields and HAS_PYPDFFORM:
                mapped = auto_map_form_fields(
                    self.form_fields, self.excel_data)
                generate_form(self.pdf_path.get(), mapped, out_path)
                mode = "form polja"
            elif self._fields:
                generate_overlay(self.pdf_path.get(), self.layout,
                                 self.excel_data, out_path)
                mode = "overlay"
            else:
                messagebox.showwarning(
                    "Upozorenje",
                    "Nema postavljenih polja. Dodajte polja s lijeve strane!")
                return
        except Exception as e:
            messagebox.showerror("Greška", f"Greška:\n{e}")
            self._status.set("Greška pri generiranju.")
            return

        self._status.set(f"PDF generiran ({mode}): {out_path}")
        if messagebox.askyesno(
                "Uspješno!",
                f"PDF generiran ({mode}):\n{out_path}\n\nOtvoriti?"):
            open_file(out_path)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    root = tk.Tk()
    HepaApp(root)
    root.mainloop()
