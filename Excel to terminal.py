import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import fitz  # PyMuPDF

def select_XLSM_file():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Select an XLSM Excel File",
        filetypes=[("Excel Files", "*.xlsm")]
    )

def select_PDF_file():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Select a PDF File",
        filetypes=[("PDF Files", "*.pdf")]
    )

def add_text_to_pdf(input_pdf, output_pdf, entries):
    pdf_document = fitz.open(input_pdf)
    try:
        for text, x, y, page_number in entries:
            if page_number >= len(pdf_document):
                print(f"Skipping invalid page {page_number}")
                continue
            page = pdf_document[page_number]
            page.insert_text((x, y), str(text), fontsize=10, color=(0, 0, 0))  # Convert to string
        pdf_document.save(output_pdf)
    finally:
        pdf_document.close()

def read_xlsm_values(file_path):
    workbook = load_workbook(file_path, keep_vba=True, data_only=True)  # Added data_only=True
    if "Podaci" not in workbook.sheetnames:
        raise ValueError("Sheet 'Podaci' not found")
    
    sheet = workbook["Podaci"]
    values = []
    
    for row in sheet.iter_rows(min_row=1, max_row=100, min_col=1, max_col=2):
        a_val = str(row[0].value).strip().lower() if row[0].value else ""
        b_val = row[1].value
        
        if b_val not in (None, ""):
            # Convert to string and clean whitespace
            clean_b = str(b_val).strip()
            values.append((a_val, clean_b))
    
    return values

try:
    xlsm_path = select_XLSM_file()
    pdf_input = select_PDF_file()
    output_file = filedialog.asksaveasfilename(
        title="Save Output PDF As",
        defaultextension=".pdf",
        filetypes=[("PDF Files", "*.pdf")]
    )

    data_pairs = read_xlsm_values(xlsm_path)
    
    # Print all values from Excel
    print("\nValues read from Excel:")
    for a, b in data_pairs:
        print(f"Column A: '{a}' | Column B: '{b}'")
    
    entries = []
    hardcoded_positions = {
        "ime investitora": [(190, 135, 0)],
        "adresa investitora (ulica)": [(140, 170, 0), (165, 460, 0)],
        "adresa investitora (grad)": [(140, 155, 0), (165, 440, 0),(120, 190, 1)],
        "oib investitora": [(450, 135, 0)],
        "lokacija građevine": [(140, 475, 0)],
        "ac snaga elektrane": [(210, 510, 0)],
        "preuzeta energija iz mreže": [(320, 545, 0)],
        "predaja u el. mrežu": [(300, 560, 0)],
        "proizvođač invertera": [(82, 680, 0)],
        "model invertera 1": [(320, 680, 0)],
        "broj omm": [(135, 760, 0)],
        "zakupljena snaga": [(480, 760, 0)]
    }

    for a_value, b_value in data_pairs:
        if a_value in hardcoded_positions:
            for x, y, page in hardcoded_positions[a_value]:
                entries.append((b_value, x, y, page))
                print(f"Mapping: {a_value} -> ({x}, {y}, page {page})")

    add_text_to_pdf(pdf_input, output_file, entries)
    print(f"\nSuccess! Modified PDF saved to: {output_file}")

except Exception as e:
    print(f"Error: {str(e)}")
